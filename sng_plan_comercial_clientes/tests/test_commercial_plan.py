# -*- coding: utf-8 -*-

import base64
from datetime import date
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from odoo.exceptions import UserError, ValidationError
from odoo.tests.common import TransactionCase


class TestCommercialPlan(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.partner = cls.env['res.partner'].create({
            'name': 'Cliente Plan A',
            'customer_rank': 1,
        })
        cls.other_partner = cls.env['res.partner'].create({
            'name': 'Cliente Plan B',
            'customer_rank': 1,
        })

    def _create_plan(self, **values):
        vals = {
            'name': 'Plan de prueba',
            'base_year': 2025,
            'target_year': 2026,
        }
        vals.update(values)
        return self.env['sng.commercial.plan'].create(vals)

    def _get_accounting_setup(self):
        income_account = self.env['account.account'].search([
            ('company_ids', 'in', self.env.company.id),
            ('account_type', '=', 'income'),
        ], limit=1)
        receivable_account = self.env['account.account'].search([
            ('company_ids', 'in', self.env.company.id),
            ('account_type', '=', 'asset_receivable'),
        ], limit=1)
        sale_journal = self.env['account.journal'].search([
            ('company_id', '=', self.env.company.id),
            ('type', '=', 'sale'),
        ], limit=1)
        bank_journal = self.env['account.journal'].search([
            ('company_id', '=', self.env.company.id),
            ('type', 'in', ('bank', 'cash')),
            ('default_account_id', '!=', False),
        ], limit=1)
        if not all((income_account, receivable_account, sale_journal, bank_journal)):
            self.skipTest('La compania de prueba no tiene configuracion contable completa.')
        self.partner.property_account_receivable_id = receivable_account
        return income_account, receivable_account, sale_journal, bank_journal

    def _create_posted_invoice(
        self,
        invoice_date,
        amount,
        move_type='out_invoice',
        partner=None,
        env=None,
        income_account=None,
        sale_journal=None,
    ):
        env = env or self.env
        partner = partner or self.partner
        if not income_account or not sale_journal:
            income_account, _, sale_journal, _ = self._get_accounting_setup()
        move = env['account.move'].create({
            'move_type': move_type,
            'partner_id': partner.id,
            'invoice_date': invoice_date,
            'journal_id': sale_journal.id,
            'invoice_line_ids': [(0, 0, {
                'name': 'Venta plan comercial',
                'account_id': income_account.id,
                'quantity': 1.0,
                'price_unit': amount,
                'tax_ids': [(5, 0, 0)],
            })],
        })
        move.action_post()
        return move

    def _pay_invoice_with_journal_entry(self, invoice, payment_date):
        _, receivable_account, _, bank_journal = self._get_accounting_setup()
        amount = invoice.amount_total
        payment_move = self.env['account.move'].create({
            'move_type': 'entry',
            'date': payment_date,
            'journal_id': bank_journal.id,
            'line_ids': [
                (0, 0, {
                    'name': 'Cobro plan comercial',
                    'account_id': bank_journal.default_account_id.id,
                    'debit': amount,
                }),
                (0, 0, {
                    'name': 'Cobro plan comercial',
                    'account_id': receivable_account.id,
                    'partner_id': self.partner.id,
                    'credit': amount,
                }),
            ],
        })
        payment_move.action_post()
        invoice_receivable = invoice.line_ids.filtered(
            lambda line: line.account_id.account_type == 'asset_receivable'
        )
        payment_receivable = payment_move.line_ids.filtered(
            lambda line: line.account_id == receivable_account
        )
        (invoice_receivable + payment_receivable).reconcile()

    def test_annual_period_and_month_progress(self):
        plan = self._create_plan()

        self.assertEqual(plan._get_base_period(), (date(2025, 1, 1), date(2025, 12, 31)))
        self.assertEqual(plan._get_target_period(), (date(2026, 1, 1), date(2026, 12, 31)))
        self.assertEqual(
            plan._get_target_period_progress(date(2026, 6, 15)),
            (12, 6, date(2026, 6, 15)),
        )
        self.assertEqual(plan._get_target_period_progress(date(2025, 12, 31)), (12, 0, False))
        self.assertEqual(
            plan._get_target_period_progress(date(2027, 1, 1)),
            (12, 12, date(2026, 12, 31)),
        )

    def test_custom_period_validation_and_partial_month_count(self):
        plan = self._create_plan(
            period_mode='custom',
            base_date_from='2025-01-15',
            base_date_to='2025-03-14',
            target_date_from='2026-04-10',
            target_date_to='2026-06-09',
        )

        self.assertEqual(plan._get_period_month_count(plan.base_date_from, plan.base_date_to), 3)
        self.assertEqual(plan._get_period_month_count(plan.target_date_from, plan.target_date_to), 3)
        self.assertEqual(
            plan._get_inclusive_period_duration(date(2024, 1, 1), date(2024, 12, 31)),
            plan._get_inclusive_period_duration(date(2025, 1, 1), date(2025, 12, 31)),
        )

        with self.assertRaises(ValidationError):
            self._create_plan(
                period_mode='custom',
                base_date_from='2025-01-01',
                base_date_to='2025-03-31',
                target_date_from='2026-01-01',
                target_date_to='2026-02-28',
            )

        with self.assertRaises(ValidationError):
            self._create_plan(
                period_mode='custom',
                base_date_from='2025-03-31',
                base_date_to='2025-01-01',
                target_date_from='2026-01-01',
                target_date_to='2026-03-31',
            )

    def test_calculation_uses_target_to_date(self):
        plan = self._create_plan()
        line = self.env['sng.commercial.plan.line'].create({
            'plan_id': plan.id,
            'partner_id': self.partner.id,
        })

        with patch.object(
            type(plan),
            '_get_sales_by_partner',
            autospec=True,
            side_effect=[{self.partner.id: 1200.0}, {self.partner.id: 300.0}],
        ), patch.object(
            type(plan),
            '_get_dpp_by_partner',
            autospec=True,
            return_value={self.partner.id: 20.0},
        ) as dpp_mock:
            plan._recalculate_plan_lines(as_of_date=date(2026, 6, 15))

        self.assertEqual(line.target_amount, 1200.0)
        self.assertEqual(line.monthly_target_amount, 100.0)
        self.assertEqual(line.target_to_date_amount, 600.0)
        self.assertEqual(line.average_monthly_sales, 50.0)
        self.assertEqual(line.projected_sales_amount, 600.0)
        self.assertEqual(line.compliance_percent, 0.5)
        self.assertEqual(line.achievement_gap_percent, -0.5)
        self.assertEqual(plan.total_target_to_date, 600.0)
        self.assertEqual(plan.total_compliance_percent, 0.5)
        dpp_mock.assert_called_once_with(
            plan,
            partner_ids=[self.partner.id],
            date_from=date(2025, 1, 1),
            date_to=date(2025, 12, 31),
        )

    def test_configuration_change_invalidates_calculation(self):
        plan = self._create_plan(state='calculated')

        plan.global_growth_factor = 1.2

        self.assertEqual(plan.state, 'draft')

    def test_closed_plan_and_lines_are_protected(self):
        plan = self._create_plan()
        line = self.env['sng.commercial.plan.line'].create({
            'plan_id': plan.id,
            'partner_id': self.partner.id,
        })
        plan.action_close()

        with self.assertRaises(UserError):
            plan.write({'name': 'Cambio no permitido'})
        with self.assertRaises(UserError):
            line.write({'notes': 'Cambio no permitido'})
        with self.assertRaises(UserError):
            line.unlink()

        plan.action_set_draft()
        plan.write({'name': 'Plan reabierto'})
        self.assertEqual(plan.name, 'Plan reabierto')

    def test_loading_customers_resynchronizes_lines(self):
        plan = self._create_plan()
        self.env['sng.commercial.plan.line'].create({
            'plan_id': plan.id,
            'partner_id': self.partner.id,
        })

        with patch.object(
            type(plan),
            '_get_sales_by_partner',
            autospec=True,
            return_value={self.other_partner.id: 250.0},
        ):
            plan.action_load_customers()

        self.assertEqual(plan.line_ids.partner_id, self.other_partner)

    def test_sales_query_filters_dates_and_credit_notes(self):
        for move_type, invoice_date, amount in (
            ('out_invoice', '2025-01-01', 100.0),
            ('out_refund', '2025-01-31', 20.0),
            ('out_invoice', '2025-02-01', 40.0),
        ):
            self._create_posted_invoice(invoice_date, amount, move_type=move_type)

        plan = self._create_plan()
        sales = plan._get_sales_by_partner(date(2025, 1, 1), date(2025, 1, 31))

        self.assertEqual(sales[self.partner.id], 80.0)

    def test_sales_consolidate_contacts_and_exclude_other_company(self):
        child = self.env['res.partner'].create({
            'name': 'Sucursal Cliente Plan A',
            'parent_id': self.partner.id,
            'type': 'invoice',
        })
        self._create_posted_invoice('2025-01-15', 75.0, partner=child)

        other_company = self.env['res.company'].create({
            'name': 'Compania Excluida Plan Comercial',
            'currency_id': self.env.company.currency_id.id,
            'payment_method_default_id': self.env['payment.methods'].search([], limit=1).id,
        })
        other_env = self.env(context={
            **self.env.context,
            'allowed_company_ids': [other_company.id],
        })
        other_income = other_env['account.account'].create({
            'name': 'Ingresos Plan Comercial Otra Compania',
            'code': 'PLANINC',
            'account_type': 'income',
            'company_ids': [(6, 0, [other_company.id])],
        })
        other_receivable = other_env['account.account'].create({
            'name': 'CxC Plan Comercial Otra Compania',
            'code': 'PLANREC',
            'account_type': 'asset_receivable',
            'reconcile': True,
            'company_ids': [(6, 0, [other_company.id])],
        })
        other_journal = other_env['account.journal'].create({
            'name': 'Ventas Plan Comercial Otra Compania',
            'code': 'PLNV',
            'type': 'sale',
            'company_id': other_company.id,
            'DA_sequence_id': other_env['ir.sequence'].create({
                'name': 'Documentos Deshabilitados Plan Comercial',
                'code': 'sequence.DA',
                'company_id': other_company.id,
            }).id,
        })
        self.partner.with_env(other_env).with_company(
            other_company
        ).property_account_receivable_id = other_receivable
        self._create_posted_invoice(
            '2025-01-15',
            500.0,
            env=other_env,
            income_account=other_income,
            sale_journal=other_journal,
        )

        plan = self._create_plan()
        sales = plan._get_sales_by_partner(date(2025, 1, 1), date(2025, 1, 31))

        self.assertEqual(sales[self.partner.id], 75.0)

    def test_dpp_filters_invoices_by_base_period(self):
        included_invoice = self._create_posted_invoice('2025-01-01', 100.0)
        excluded_invoice = self._create_posted_invoice('2025-02-01', 100.0)
        self._pay_invoice_with_journal_entry(included_invoice, '2025-01-11')
        self._pay_invoice_with_journal_entry(excluded_invoice, '2025-02-21')

        metric_model = self.env['regalarte.customer.metric']
        historical_dpp = metric_model._get_dpp_metrics(
            [self.partner.id],
            self.env.company,
        )
        filtered_dpp = metric_model._get_dpp_metrics(
            [self.partner.id],
            self.env.company,
            date_from=date(2025, 1, 1),
            date_to=date(2025, 1, 31),
        )
        excluded_period_dpp = metric_model._get_dpp_metrics(
            [self.partner.id],
            self.env.company,
            date_from=date(2025, 2, 1),
            date_to=date(2025, 2, 28),
        )

        self.assertNotEqual(
            filtered_dpp[self.partner.id],
            excluded_period_dpp[self.partner.id],
        )
        self.assertAlmostEqual(
            historical_dpp[self.partner.id],
            (filtered_dpp[self.partner.id] + excluded_period_dpp[self.partner.id]) / 2.0,
        )

    def test_xlsx_has_separate_compliance_and_deviation(self):
        plan = self._create_plan()
        self.env['sng.commercial.plan.line'].create({
            'plan_id': plan.id,
            'partner_id': self.partner.id,
            'target_amount': 100.0,
            'target_to_date_amount': 50.0,
            'current_sales_amount': 40.0,
            'compliance_percent': 0.8,
            'achievement_gap_percent': -0.2,
        })

        _, encoded_file = plan.generate_xlsx_file()
        workbook = load_workbook(BytesIO(base64.b64decode(encoded_file)), data_only=True)
        worksheet = workbook['Plan Comercial']

        self.assertEqual(worksheet.cell(6, 17).value, 'Cumplimiento')
        self.assertEqual(worksheet.cell(6, 18).value, 'Desviacion vs Meta')
        self.assertEqual(worksheet.cell(7, 17).value, 0.8)
        self.assertEqual(worksheet.cell(7, 18).value, -0.2)
