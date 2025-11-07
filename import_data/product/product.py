# -*- coding: utf-8 -*-
# Ejecutar en Odoo Shell:  odoo-bin shell -d <tu_db>

import os
import re
import openpyxl

PRODUCT = env['product.template']
CATEGORY = env['product.category']
CABYS_PRODUCT = env['cabys.producto']  # Modelo CAByS
PRICELIST = env['product.pricelist']
PRICELIST_ITEM = env['product.pricelist.item']

# Contexto para evitar tracking / chatter
CTX_NOTRACK = {
    'tracking_disable': True,
    'mail_create_nolog': True,
    'mail_notrack': True,
}
PRODUCT_NOTRACK = PRODUCT.with_context(**CTX_NOTRACK)
CATEGORY_NOTRACK = CATEGORY.with_context(**CTX_NOTRACK)
PRICELIST_NOTRACK = PRICELIST.with_context(**CTX_NOTRACK)
PLI_NOTRACK = PRICELIST_ITEM.with_context(**CTX_NOTRACK)

# ===== Config =====
NOMBRE_HOJA = 'Productos'           # cambia si tu hoja se llama distinto
COMMIT_INTERVAL = 50                # commit cada N filas
DEFAULT_XLSX = 'inventario.xlsx'    # nombre por defecto

# ===== Utils =====
def _parse_number(val):
    """Convierte '1,098.41' | '1.098,41' | 1098.41 | '' -> float o None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.upper() in {'NA', 'N/A', 'NONE', 'NULL', '-'}:
        return None
    s = s.replace(' ', '')
    if ',' in s and '.' in s:
        # 1.098,41 -> 1098.41 ; 1,098.41 -> 1098.41
        if re.search(r'^\d{1,3}(\.\d{3})+,\d{1,2}$', s):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s and '.' not in s:
        s = s.replace(',', '.') if s.count(',') == 1 else s.replace(',', '')
    try:
        return float(s)
    except Exception:
        return None


def _parse_bool(val):
    if isinstance(val, bool):
        return val
    s = str(val or '').strip().lower()
    return s in {'true', '1', 'si', 's√≠', 'yes', 'y', 't', 'verdadero'}


def _normalize_header(s):
    return re.sub(r'\s+', ' ', str(s or '').strip().lower())


def _find_field(model, candidates):
    fields = env[model]._fields
    for c in candidates:
        if c in fields:
            return c
    return None


def _get_or_create_categ_by_path(path):
    """
    Crea la jerarqu√≠a si no existe. Acepta 'Padre / Hija / Nieta'.
    Devuelve record de product.category o None.
    """
    parts = [p.strip() for p in str(path or '').split('/') if str(p).strip()]
    if not parts:
        return None
    parent = None
    for name in parts:
        dom = [('name', '=', name), ('parent_id', '=', parent.id if parent else False)]
        rec = CATEGORY.search(dom, limit=1)
        if not rec:
            rec = CATEGORY_NOTRACK.create({'name': name, 'parent_id': parent.id if parent else False})
        parent = rec
    return parent


def _resolve_path(filename):
    """Devuelve ruta existente. Prioriza carpeta del script (__file__), luego CWD, luego carpeta padre."""
    if os.path.isabs(filename):
        if not os.path.exists(filename):
            raise FileNotFoundError(f"No se encontr√≥ el archivo absoluto: {filename}")
        return filename
    tried = []
    base_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    p1 = os.path.join(base_dir, filename); tried.append(p1)
    p2 = os.path.join(os.getcwd(), filename); tried.append(p2)
    p3 = os.path.join(os.path.dirname(base_dir), filename); tried.append(p3)
    for p in tried:
        if os.path.exists(p):
            return p
    raise FileNotFoundError("No se encontr√≥ el archivo '{0}'. Prob√©:\n- {1}".format(filename, "\n- ".join(tried)))


# ===== Input archivo =====
ARCHIVO_IN = input(f"üìÑ Nombre del XLSX (enter = {DEFAULT_XLSX}): ").strip() or DEFAULT_XLSX
ARCHIVO = _resolve_path(ARCHIVO_IN)
print(f"üìÇ Usando archivo: {ARCHIVO}")

# ===== Cargar libro / hoja =====
wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
print(f"üìë Hojas disponibles: {wb.sheetnames}")
if NOMBRE_HOJA not in wb.sheetnames:
    raise ValueError(f"La hoja '{NOMBRE_HOJA}' no existe. Hojas: {wb.sheetnames}")
ws = wb[NOMBRE_HOJA]

# ===== Mapeo encabezados =====
# OJO: los keys de este dict son el resultado de _normalize_header sobre tu Excel.
header_map = {
    # COD_PROD -> ref interna (solo para llenar, NO para buscar)
    'cod_prod': 'default_code',

    # NOMBRE_1 -> nombre de producto
    'nombre_1': 'name',

    # COSTO_FINAL -> costo (standard_price)
    'costo_final': 'standard_price',

    # PRECIO_3 (precio publico en colon) -> precio de venta
    'precio_3 (precio publico en colon)': 'list_price',
    'precio_3': 'list_price',  # por si cambias el header a algo m√°s corto

    # PRECIO_1 (Mayor) -> lista de precios "Mayor"
    'precio_1 (mayor)': 'price_mayor',
    'precio_1': 'price_mayor',

    # PRECIO_9 (mayor en dolares) -> lista de precios "Mayor d√≥lares"
    'precio_9 (mayor en dolares)': 'price_mayor_usd',
    'precio_9': 'price_mayor_usd',

    # Categor√≠as: DES_FAMI = padre, DES_SUBF = subcategor√≠a
    'des_fami': 'categ_parent',
    'des_subf': 'categ_child',

    # CAByS
    'cod_cabys': 'cabys_code',

    # Flags opcionales (si los tuvieras en el Excel futuro)
    'se puede comprar': 'purchase_ok',
    'se puede vender': 'sale_ok',
}

rows = list(ws.iter_rows(values_only=True))
if not rows:
    raise ValueError("La hoja est√° vac√≠a.")

headers = rows[0]
norm_headers = [_normalize_header(h) for h in headers]
idx = {}
for i, h in enumerate(norm_headers):
    if h in header_map:
        idx[header_map[h]] = i

# Verifica campo Many2one en product.template
has_cabys_m2o = _find_field('product.template', ['cabys_product_id']) is not None


def _getv(row, key):
    col = idx.get(key, None)
    return row[col] if (col is not None and col < len(row)) else None


# ===== Crear/obtener listas de precios =====
pl_mayor = PRICELIST.search([('name', '=', 'Mayor')], limit=1)
if not pl_mayor:
    pl_mayor = PRICELIST_NOTRACK.create({'name': 'Mayor'})
    print("üßæ Creada lista de precios 'Mayor'")

pl_mayor_usd = False
cur_usd = env['res.currency'].search([('name', '=', 'USD')], limit=1)
if not cur_usd:
    print("‚ö† No se encontr√≥ moneda USD; no se crear√° la lista de precios 'Mayor d√≥lares'")
else:
    pl_mayor_usd = PRICELIST.search([('name', '=', 'Mayor d√≥lares')], limit=1)
    if not pl_mayor_usd:
        pl_mayor_usd = PRICELIST_NOTRACK.create({
            'name': 'Mayor d√≥lares',
            'currency_id': cur_usd.id,
        })
        print("üßæ Creada lista de precios 'Mayor d√≥lares' (USD)")

# ===== Limpiar items existentes de esas listas =====
if pl_mayor:
    old_items = PLI_NOTRACK.search([('pricelist_id', '=', pl_mayor.id)])
    if old_items:
        print(f"üóë Eliminando {len(old_items)} items de la lista 'Mayor'...")
        old_items.unlink()

if pl_mayor_usd:
    old_items2 = PLI_NOTRACK.search([('pricelist_id', '=', pl_mayor_usd.id)])
    if old_items2:
        print(f"üóë Eliminando {len(old_items2)} items de la lista 'Mayor d√≥lares'...")
        old_items2.unlink()

# ===== Procesamiento =====
errores = []
logs = []
procesados = 0
created_total = 0
updated_total = 0
skipped_no_keys = 0
skipped_multi = 0
commit_counter = 0

for r, row in enumerate(rows[1:], start=2):  # desde fila 2 (despu√©s de encabezados)
    procesados += 1
    try:
        # Claves b√°sicas
        name = _getv(row, 'name')
        name = str(name).strip() if name is not None else ''
        ref = _getv(row, 'default_code')
        ref = str(ref).strip() if ref is not None else ''

        if not name:
            skipped_no_keys += 1
            logs.append(f"[{r}] sin 'Nombre' -> saltado")
            continue

        # Precios para listas de precios
        price_mayor = _parse_number(_getv(row, 'price_mayor'))
        price_mayor_usd = _parse_number(_getv(row, 'price_mayor_usd'))

        # ===== Categor√≠a: DES_FAMI (padre) + DES_SUBF (hija) =====
        categ = None
        parent_name = _getv(row, 'categ_parent')
        child_name = _getv(row, 'categ_child')
        parent_name = str(parent_name).strip() if parent_name else ''
        child_name = str(child_name).strip() if child_name else ''

        path = None
        if parent_name and child_name:
            path = f"{parent_name} / {child_name}"
        elif parent_name:
            path = parent_name
        elif child_name:
            path = child_name

        if path:
            categ = _get_or_create_categ_by_path(path)

        # ===== Buscar existente por (name, categ_id) =====
        prod = False
        domain = [('name', '=', name)]
        if categ:
            domain.append(('categ_id', '=', categ.id))
        found = PRODUCT.search(domain)

        if len(found) == 1:
            prod = found[0]
        elif len(found) > 1:
            skipped_multi += 1
            logs.append(
                f"[{r}] NAME='{name}', CATEG='{categ.display_name if categ else '-'}' "
                f"m√∫ltiples templates {found.ids} -> saltado"
            )
            continue

        # --- Valores comunes (create/write) ---
        vals = {}

        # Nombre y referencia interna
        vals['name'] = name
        if ref:
            vals['default_code'] = ref

        # Categor√≠a en vals
        if categ:
            vals['categ_id'] = categ.id

        # Precios base (list_price / standard_price)
        lp = _parse_number(_getv(row, 'list_price'))
        if lp is not None:
            vals['list_price'] = lp

        sp = _parse_number(_getv(row, 'standard_price'))
        if sp is not None:
            vals['standard_price'] = sp

        # CAByS como Many2one (cabys_product_id -> cabys.producto.codigo)
        cabys_val = _getv(row, 'cabys_code')
        cabys_val = str(cabys_val).strip() if cabys_val is not None else ''
        if has_cabys_m2o and cabys_val:
            cabys_rec = CABYS_PRODUCT.with_context(active_test=False).search([('codigo', '=', cabys_val)], limit=1)
            if cabys_rec:
                vals['cabys_product_id'] = cabys_rec.id
            else:
                logs.append(f"[{r}] CAByS '{cabys_val}' no encontrado en cabys.producto -> no se asigna")

        # Flags (solo si existen esas columnas en el Excel)
        if 'purchase_ok' in idx:
            vals['purchase_ok'] = _parse_bool(_getv(row, 'purchase_ok'))
        if 'sale_ok' in idx:
            vals['sale_ok'] = _parse_bool(_getv(row, 'sale_ok'))

        # --- Crear o actualizar SIN tracking ---
        if prod:
            PRODUCT_NOTRACK.browse(prod.id).write(vals)
            updated_total += 1
            print(f"[{r}] ‚úÖ ACTUALIZADO name='{name}' ref={ref or '-'} -> {vals}")
        else:
            vals.setdefault('type', 'consu')
            vals.setdefault('purchase_ok', True)
            vals.setdefault('sale_ok', True)
            prod = PRODUCT_NOTRACK.create(vals)
            created_total += 1
            print(f"[{r}] üÜï CREADO name='{name}' ref={ref or '-'} -> {vals}")

        # ===== Actualizar / crear items de lista de precios =====

        # Lista "Mayor" (colones) -> solo si price_mayor > 0
        if pl_mayor and price_mayor is not None and price_mayor > 0:
            item = PLI_NOTRACK.search([
                ('pricelist_id', '=', pl_mayor.id),
                ('product_tmpl_id', '=', prod.id),
                ('display_applied_on', '=', '1_product'),
            ], limit=1)
            vals_item = {
                'pricelist_id': pl_mayor.id,
                'product_tmpl_id': prod.id,
                'display_applied_on': '1_product',
                'compute_price': 'fixed',
                'fixed_price': price_mayor,
            }
            if item:
                item.write(vals_item)
            else:
                PLI_NOTRACK.create(vals_item)

        # Lista "Mayor d√≥lares" (USD) -> solo si price_mayor_usd > 0
        if pl_mayor_usd and price_mayor_usd is not None and price_mayor_usd > 0:
            item2 = PLI_NOTRACK.search([
                ('pricelist_id', '=', pl_mayor_usd.id),
                ('product_tmpl_id', '=', prod.id),
                ('display_applied_on', '=', '1_product'),
            ], limit=1)
            vals_item2 = {
                'pricelist_id': pl_mayor_usd.id,
                'product_tmpl_id': prod.id,
                'display_applied_on': '1_product',
                'compute_price': 'fixed',
                'fixed_price': price_mayor_usd,
            }
            if item2:
                item2.write(vals_item2)
            else:
                PLI_NOTRACK.create(vals_item2)

    except Exception as e:
        env.cr.rollback()
        errores.append(f"[{r}] ‚ùå Error name='{name}' ref={ref or '-'} -> {e}")
        print(errores[-1])

    # Commits por lotes
    commit_counter += 1
    if commit_counter >= COMMIT_INTERVAL:
        env.cr.commit()
        commit_counter = 0

# Commit final
if commit_counter > 0:
    env.cr.commit()

# ===== Logs / Errores =====
base_dir = os.getcwd()
ruta_errores = os.path.join(base_dir, 'errores_importacion.txt')
ruta_log = os.path.join(base_dir, 'import_log.txt')

if errores:
    with open(ruta_errores, 'w', encoding='utf-8') as f:
        f.write('\n'.join(errores))
    print(f"‚ö†Ô∏è {len(errores)} errores guardados en: {ruta_errores}")

with open(ruta_log, 'w', encoding='utf-8') as f:
    f.write('\n'.join(logs))
print(f"üìù Log de proceso en: {ruta_log}")

# ===== Resumen =====
print("=== RESUMEN ===")
print(f"Procesados: {procesados}")
print(f"Creados: {created_total}")
print(f"Actualizados: {updated_total}")
print(f"Saltados sin claves (nombre): {skipped_no_keys}")
print(f"Saltados por m√∫ltiples coincidencias (nombre+categ): {skipped_multi}")
print(f"Errores: {len(errores)}")
