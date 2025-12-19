import os
import re

try:
    import openpyxl
except ImportError:
    raise Exception("Falta openpyxl. En Odoo.sh normalmente está disponible. Si no, usa CSV en su lugar.")

PARTNER = env['res.partner'].sudo()
PAYTERM = env['account.payment.term'].sudo()
COUNTRY = env['res.country'].sudo()

# ======================
# Config
# ======================
DEFAULT_XLSX = "proveedores.xlsx"
SHEET_NAME = None  # None = primera hoja
COMMIT_INTERVAL = 50

# Columnas esperadas (header exacto o parecido; se normaliza)
COLS = {
    "nombre": "name",
    "cédula": "vat",
    "cedula": "vat",
    "país": "country",
    "pais": "country",
    "teléfono": "phone",
    "telefono": "phone",
    "plazo": "plazo",
    "dirección": "street",
    "direccion": "street",
}

# Contexto para importación (reduce tracking y recomputes ruidosos)
CTX = {
    "tracking_disable": True,
    "mail_notrack": True,
    "mail_create_nosubscribe": True,
    # Algunos módulos revisan estas banderas; si no existen, no pasa nada:
    "no_vat_validation": True,
    "skip_peppol_validation": True,
}

# ======================
# Utils
# ======================
def _normalize_header(s):
    s = str(s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _digits_only(s):
    """Deja solo dígitos (para cédula)."""
    return re.sub(r"\D+", "", str(s or "").strip())

def _clean_str(s):
    return str(s or "").strip()

def _to_int(s, default=0):
    s = _clean_str(s)
    if not s:
        return default
    # soporta "30", "30.0", etc.
    if re.match(r"^\d+(\.0+)?$", s):
        return int(float(s))
    d = _digits_only(s)
    return int(d) if d else default

def _resolve_path(filename):
    """Busca el archivo en ubicaciones típicas."""
    if os.path.isabs(filename) and os.path.exists(filename):
        return filename

    tried = []
    base_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
    candidates = [
        os.path.join(base_dir, filename),
        os.path.join(os.getcwd(), filename),
        os.path.join("/home/odoo", filename),
        os.path.join("/home/odoo/src", filename),
        os.path.join("/home/odoo/src/user", filename),
    ]
    for p in candidates:
        tried.append(p)
        if os.path.exists(p):
            return p
    raise FileNotFoundError("No se encontró el archivo. Probé:\n- " + "\n- ".join(tried))

def _get_country_id(code):
    code = _clean_str(code).upper()
    if not code:
        return False
    country = COUNTRY.search([("code", "=", code)], limit=1)
    return country.id if country else False

def _find_payment_term(plazo_days):
    """Busca un payment term existente que represente el plazo (por nombre)."""
    if not plazo_days:
        return False
    term = PAYTERM.search([("name", "ilike", str(plazo_days))], limit=1)
    return term.id if term else False


# ======================
# Input
# ======================
file_in = input(f"📄 Nombre del XLSX (enter = {DEFAULT_XLSX}): ").strip() or DEFAULT_XLSX
path = _resolve_path(file_in)
print(f"📂 Usando archivo: {path}")

wb = openpyxl.load_workbook(path, data_only=True)
ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

# ======================
# Headers -> índices
# ======================
header_row = None
headers = None

for r in range(1, 6):  # intenta primeras 5 filas como header
    values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    norm = [_normalize_header(v) for v in values]
    print(f"Fila {r}: {norm}")
    if any(h in COLS for h in norm):
        header_row = r
        headers = norm
        break

if not header_row:
    raise Exception("No pude detectar encabezados. Asegúrate que exista: Nombre, Cédula, País, Teléfono, Plazo, Dirección.")

idx = {}
for i, h in enumerate(headers, start=1):
    if h in COLS:
        idx[COLS[h]] = i

required = ["name", "vat"]
missing = [k for k in required if k not in idx]
if missing:
    raise Exception(f"Faltan columnas obligatorias: {missing}. Detectadas: {list(idx.keys())}")

# ======================
# Proceso
# ======================
procesados = 0
creados = 0
actualizados = 0
errores = []
skips = 0
commit_counter = 0

print(f"🚀 Iniciando importación desde fila {header_row + 1}...")

for r in range(header_row + 1, ws.max_row + 1):
    procesados += 1

    try:
        name = _clean_str(ws.cell(row=r, column=idx["name"]).value)
        vat_raw = ws.cell(row=r, column=idx["vat"]).value
        vat = _digits_only(vat_raw)

        country_code = _clean_str(ws.cell(row=r, column=idx.get("country", 0)).value) if idx.get("country") else ""
        phone = _clean_str(ws.cell(row=r, column=idx.get("phone", 0)).value) if idx.get("phone") else ""
        street = _clean_str(ws.cell(row=r, column=idx.get("street", 0)).value) if idx.get("street") else ""
        plazo = _to_int(ws.cell(row=r, column=idx.get("plazo", 0)).value) if idx.get("plazo") else 0

        # Validaciones mínimas
        if not name and not vat:
            skips += 1
            continue

        # ======================
        # Buscar existente:
        # 1) por vat si existe
        # 2) fallback por name + supplier_rank>0
        # ======================
        partner = False

        if vat and len(vat) >= 6:
            found = PARTNER.search([("vat", "=", vat)], limit=2)
            if len(found) == 1:
                partner = found

        if not partner and name:
            found = PARTNER.search([("name", "=", name), ("supplier_rank", ">", 0)], limit=2)
            if len(found) == 1:
                partner = found

        # ======================
        # Construcción de vals
        # ======================
        vals = {
            "supplier_rank": 1,  # asegura proveedor
        }

        if name:
            vals["name"] = name
        if vat:
            vals["vat"] = vat

        country_id = _get_country_id(country_code)
        if country_id:
            vals["country_id"] = country_id
        if phone:
            vals["phone"] = phone
        if street:
            vals["street"] = street

        term_id = _find_payment_term(plazo)
        if term_id:
            vals["property_payment_term_id"] = term_id

        # Si se crea, por defecto lo ponemos como empresa (ajusta si quieres)
        if not partner:
            vals.setdefault("company_type", "company")

        # ======================
        # Crear / actualizar
        # ======================
        if partner:
            partner.with_context(**CTX).write(vals)
            actualizados += 1
        else:
            PARTNER.with_context(**CTX).create(vals)
            creados += 1

    except Exception as e:
        # ✅ CLAVE: rollback por fila para limpiar la transacción abortada
        env.cr.rollback()
        errores.append(f"Fila {r}: {type(e).__name__}: {e}")
        print(f"❌ Error en fila {r}: {type(e).__name__}: {e}")
        continue

    # Commit por lotes (solo si la fila no falló)
    commit_counter += 1
    if commit_counter >= COMMIT_INTERVAL:
        try:
            env.cr.commit()
        except Exception as e:
            # Si por algún recompute raro el commit falla, rollback y seguimos
            env.cr.rollback()
            errores.append(f"COMMIT lote en fila {r}: {type(e).__name__}: {e}")
            print(f"❌ Error en COMMIT (fila {r}): {type(e).__name__}: {e}")
        commit_counter = 0

# Commit final
try:
    env.cr.commit()
except Exception as e:
    env.cr.rollback()
    errores.append(f"COMMIT final: {type(e).__name__}: {e}")
    print(f"❌ Error en COMMIT final: {type(e).__name__}: {e}")

print("\n📊 RESUMEN PROVEEDORES")
print(f"Procesados:   {procesados}")
print(f"Creados:      {creados}")
print(f"Actualizados: {actualizados}")
print(f"Saltados:     {skips}")
print(f"Errores:      {len(errores)}")

if errores:
    out_path = os.path.join(os.getcwd(), "errores_import_proveedores.txt")
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("\n".join(errores))
        print(f"⚠️ Errores guardados en: {out_path}")
    except Exception:
        print("⚠️ No pude guardar archivo de errores. Aquí van los primeros 30:")
        for line in errores[:30]:
            print(" -", line)
