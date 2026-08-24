# -*- coding: utf-8 -*-
from io import BytesIO

from openpyxl import load_workbook

from odoo import fields
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.tests.common import TransactionCase, new_test_user


class TestCycleCountApproval(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.company.email = "cycle-count@example.com;secondary@example.com"
        cls.warehouse = cls.env["stock.warehouse"].search(
            [("company_id", "=", cls.company.id)],
            limit=1,
        )
        cls.location = cls.warehouse.lot_stock_id
        cls.product_gain = cls.env["product.product"].create(
            {
                "name": "Cycle Count Gain",
                "is_storable": True,
                "standard_price": 10.0,
            }
        )
        cls.product_shortage = cls.env["product.product"].create(
            {
                "name": "Cycle Count Shortage",
                "is_storable": True,
                "standard_price": 20.0,
            }
        )
        cls.env["stock.quant"]._update_available_quantity(
            cls.product_gain,
            cls.location,
            10.0,
        )
        cls.env["stock.quant"]._update_available_quantity(
            cls.product_shortage,
            cls.location,
            5.0,
        )
        cls.quant_gain = cls.env["stock.quant"].search(
            [
                ("product_id", "=", cls.product_gain.id),
                ("location_id", "=", cls.location.id),
            ],
            limit=1,
        )
        cls.quant_shortage = cls.env["stock.quant"].search(
            [
                ("product_id", "=", cls.product_shortage.id),
                ("location_id", "=", cls.location.id),
            ],
            limit=1,
        )
        cls.config = cls.env["sng.cycle.count.config"].create(
            {
                "name": "Cycle Count Test",
                "company_id": cls.company.id,
                "selection_method": "random",
                "daily_product_count": 2,
                "location_ids": [(6, 0, cls.location.ids)],
            }
        )
        cls.operator = new_test_user(
            cls.env,
            login="cycle.operator@example.com",
            groups="sng_cycle_count.group_cycle_count_operator",
            company_id=cls.company.id,
            name="Cycle Operator",
        )
        cls.supervisor = new_test_user(
            cls.env,
            login="cycle.supervisor@example.com",
            groups="sng_cycle_count.group_cycle_count_supervisor",
            company_id=cls.company.id,
            name="Cycle Supervisor",
        )
        cls.manager = new_test_user(
            cls.env,
            login="cycle.manager@example.com",
            groups="sng_cycle_count.group_cycle_count_management",
            company_id=cls.company.id,
            name="Cycle Manager",
        )

    def _create_count(self, include_shortage=False):
        count = self.env["sng.cycle.count"].create(
            {
                "count_date": fields.Date.today(),
                "config_id": self.config.id,
                "user_id": self.operator.id,
                "company_id": self.company.id,
            }
        )
        line_values = [
            {
                "cycle_count_id": count.id,
                "quant_id": self.quant_gain.id,
                "theoretical_qty": self.quant_gain.quantity,
            }
        ]
        if include_shortage:
            line_values.append(
                {
                    "cycle_count_id": count.id,
                    "quant_id": self.quant_shortage.id,
                    "theoretical_qty": self.quant_shortage.quantity,
                }
            )
        self.env["sng.cycle.count.line"].create(line_values)
        return count

    def _disable_quantity_reports(self):
        model_class = self.env.registry["sng.cycle.count"]
        self.patch(model_class, "_generate_discrepancy_reports", lambda records: True)

    def test_pending_lines_are_not_discrepancies(self):
        count = self._create_count()

        self.assertEqual(count.counted_lines, 0)
        self.assertEqual(count.discrepancy_lines, 0)
        self.assertFalse(count.has_discrepancies)
        self.assertEqual(count.line_ids.difference_qty, 0.0)
        self.assertEqual(count.line_ids.difference_value, 0.0)
        self.assertEqual(count.total_theoretical_value, 100.0)
        self.assertEqual(count.total_counted_value, 0.0)

    def test_cost_is_frozen_and_totals_separate_gain_and_shortage(self):
        count = self._create_count(include_shortage=True)
        gain_line = count.line_ids.filtered(lambda line: line.product_id == self.product_gain)
        shortage_line = count.line_ids.filtered(
            lambda line: line.product_id == self.product_shortage
        )
        self.product_gain.standard_price = 99.0

        gain_line.with_user(self.operator).write(
            {"counted_qty": 12.0, "state": "counted", "count_date": fields.Datetime.now()}
        )
        shortage_line.with_user(self.operator).write(
            {"counted_qty": 4.0, "state": "counted", "count_date": fields.Datetime.now()}
        )

        self.assertEqual(gain_line.unit_cost, 10.0)
        self.assertEqual(shortage_line.unit_cost, 20.0)
        self.assertEqual(count.total_theoretical_value, 200.0)
        self.assertEqual(count.total_counted_value, 200.0)
        self.assertEqual(count.total_gain_value, 20.0)
        self.assertEqual(count.total_shortage_value, 20.0)
        self.assertEqual(count.total_difference_value, 0.0)
        self.assertEqual(count.discrepancy_lines, 2)

    def test_submit_requires_all_lines_and_does_not_adjust_stock(self):
        count = self._create_count()
        with self.assertRaises(UserError):
            count.with_user(self.operator).action_submit_for_approval()

        self._disable_quantity_reports()
        count.line_ids.with_user(self.operator).write(
            {"counted_qty": 8.0, "state": "counted", "count_date": fields.Datetime.now()}
        )
        count.with_user(self.operator).action_submit_for_approval()

        with self.assertRaises(UserError):
            self.env["sng.cycle.count.line"].with_user(self.operator).create(
                {
                    "cycle_count_id": count.id,
                    "quant_id": self.quant_shortage.id,
                    "theoretical_qty": self.quant_shortage.quantity,
                }
            )

        self.quant_gain.invalidate_recordset(["quantity"])
        self.assertEqual(count.state, "pending_approval")
        self.assertEqual(self.quant_gain.quantity, 10.0)
        self.assertEqual(count.submitted_by_id, self.operator)
        approval_type = self.env.ref("sng_cycle_count.mail_activity_cycle_count_approval")
        activity = self.env["mail.activity"].search(
            [
                ("res_model", "=", count._name),
                ("res_id", "=", count.id),
                ("activity_type_id", "=", approval_type.id),
                ("user_id", "=", self.manager.id),
            ]
        )
        self.assertEqual(len(activity), 1)
        mail = self.env["mail.mail"].search(
            [("email_to", "=", self.manager.email), ("subject", "ilike", count.name)]
        )
        self.assertEqual(len(mail), 1)
        self.assertEqual(mail.state, "outgoing")
        self.assertEqual(mail.email_from, "cycle-count@example.com")

    def test_only_management_can_approve_and_adjustment_is_atomic(self):
        self._disable_quantity_reports()
        count = self._create_count()
        count.line_ids.with_user(self.operator).write(
            {"counted_qty": 8.0, "state": "counted", "count_date": fields.Datetime.now()}
        )
        count.with_user(self.operator).action_submit_for_approval()

        with self.assertRaises(AccessError):
            count.with_user(self.operator).action_approve()
        with self.assertRaises(AccessError):
            count.with_user(self.supervisor).action_approve()
        with self.assertRaises(UserError):
            count.line_ids.with_user(self.operator).write({"counted_qty": 7.0})

        count.with_user(self.manager).action_approve()
        self.quant_gain.invalidate_recordset(["quantity"])
        self.assertEqual(count.state, "done")
        self.assertEqual(count.line_ids.state, "adjusted")
        self.assertEqual(self.quant_gain.quantity, 8.0)
        self.assertEqual(count.approved_by_id, self.manager)
        self.assertTrue(count.approved_at)
        with self.assertRaises(UserError):
            count.action_reopen()

    def test_stock_change_blocks_approval_and_return_resets_changed_lines(self):
        self._disable_quantity_reports()
        count = self._create_count(include_shortage=True)
        count.line_ids.with_user(self.operator).write(
            {"state": "counted", "count_date": fields.Datetime.now()}
        )
        count.line_ids.filtered(lambda line: line.product_id == self.product_gain).with_user(
            self.operator
        ).write({"counted_qty": 9.0})
        count.line_ids.filtered(
            lambda line: line.product_id == self.product_shortage
        ).with_user(self.operator).write({"counted_qty": 5.0})
        count.with_user(self.operator).action_submit_for_approval()

        self.env["stock.quant"]._update_available_quantity(
            self.product_gain,
            self.location,
            1.0,
        )
        with self.assertRaises(UserError):
            count.with_user(self.manager).action_approve()
        self.assertEqual(count.state, "pending_approval")

        wizard = self.env["sng.cycle.count.return.wizard"].with_user(self.manager).create(
            {"cycle_count_id": count.id, "reason": "El stock cambió durante la revisión."}
        )
        wizard.action_confirm()

        gain_line = count.line_ids.filtered(lambda line: line.product_id == self.product_gain)
        unchanged_line = count.line_ids.filtered(
            lambda line: line.product_id == self.product_shortage
        )
        self.assertEqual(count.state, "in_progress")
        self.assertEqual(gain_line.state, "pending")
        self.assertEqual(gain_line.theoretical_qty, 11.0)
        self.assertEqual(gain_line.counted_qty, 0.0)
        self.assertEqual(unchanged_line.state, "counted")
        self.assertEqual(unchanged_line.counted_qty, 5.0)
        recount_type = self.env.ref("sng_cycle_count.mail_activity_cycle_count_recount")
        self.assertEqual(
            self.env["mail.activity"].search_count(
                [
                    ("res_model", "=", count._name),
                    ("res_id", "=", count.id),
                    ("activity_type_id", "=", recount_type.id),
                    ("user_id", "=", self.operator.id),
                ]
            ),
            1,
        )

    def test_server_guards_and_cost_visibility(self):
        count = self._create_count()
        with self.assertRaises(AccessError):
            count.with_user(self.operator).write({"state": "pending_approval"})
        with self.assertRaises(ValidationError), self.env.cr.savepoint():
            count.line_ids.with_user(self.operator).write({"counted_qty": -1.0})
        with self.assertRaises(ValidationError), self.env.cr.savepoint():
            self.env["sng.cycle.count.line"].create(
                {
                    "cycle_count_id": count.id,
                    "quant_id": self.quant_gain.id,
                    "theoretical_qty": self.quant_gain.quantity,
                }
            )
        with self.assertRaises(AccessError):
            count.line_ids.with_user(self.operator).read(["unit_cost"])

        supervisor_values = count.line_ids.with_user(self.supervisor).read(["unit_cost"])
        self.assertEqual(supervisor_values[0]["unit_cost"], 10.0)

    def test_resubmission_replaces_management_activity(self):
        count = self._create_count()
        count.line_ids.with_user(self.operator).action_copy_theoretical()
        count.with_user(self.operator).action_submit_for_approval()

        wizard = self.env["sng.cycle.count.return.wizard"].with_user(self.manager).create(
            {"cycle_count_id": count.id, "reason": "Confirmar nuevamente el conteo."}
        )
        wizard.action_confirm()
        count.with_user(self.operator).action_submit_for_approval()

        approval_type = self.env.ref("sng_cycle_count.mail_activity_cycle_count_approval")
        self.assertEqual(
            self.env["mail.activity"].search_count(
                [
                    ("res_model", "=", count._name),
                    ("res_id", "=", count.id),
                    ("activity_type_id", "=", approval_type.id),
                    ("user_id", "=", self.manager.id),
                ]
            ),
            1,
        )

    def test_valued_reports_are_restricted_and_render(self):
        count = self._create_count()
        count.line_ids.with_user(self.operator).write(
            {"counted_qty": 8.0, "state": "counted", "count_date": fields.Datetime.now()}
        )

        report_model = self.env["ir.actions.report"]
        html, html_type = report_model.with_user(self.supervisor)._render_qweb_html(
            "sng_cycle_count.action_report_cycle_count_valuation",
            count.ids,
        )
        xlsx, xlsx_type = report_model.with_user(self.supervisor)._render_xlsx(
            "sng_cycle_count.action_report_cycle_count_valuation_xlsx",
            count.ids,
            {},
        )
        workbook = load_workbook(BytesIO(xlsx), read_only=True)

        self.assertEqual(html_type, "html")
        self.assertIn(b"Reporte Valorizado", html)
        self.assertEqual(xlsx_type, "xlsx")
        self.assertEqual(workbook.active.cell(8, 1).value, "Producto")
        with self.assertRaises(AccessError):
            report_model.with_user(self.operator)._render_qweb_html(
                "sng_cycle_count.action_report_cycle_count_valuation",
                count.ids,
            )
        with self.assertRaises(AccessError):
            report_model.with_user(self.operator)._render_xlsx(
                "sng_cycle_count.action_report_cycle_count_valuation_xlsx",
                count.ids,
                {},
            )
