# -*- coding: utf-8 -*-

from io import BytesIO
from unittest import SkipTest

from openpyxl import load_workbook

from odoo import Command
from odoo.exceptions import UserError, ValidationError
from odoo.tests.common import TransactionCase, new_test_user


class TestCustomerProductInvoiceReport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.env = cls.env(context=dict(cls.env.context, tracking_disable=True))
        if "payment_method_default_id" in cls.env.company._fields:
            payment_method = cls.env.ref(
                "cr_electronic_invoice.PaymentMethods_1",
                raise_if_not_found=False,
            )
            if payment_method:
                cls.env.company.payment_method_default_id = payment_method
        cls.income_account = cls.env["account.account"].search(
            [
                ("account_type", "=", "income"),
                ("company_ids", "in", cls.env.company.id),
                ("deprecated", "=", False),
            ],
            limit=1,
        )
        cls.receivable_account = cls.env["account.account"].search(
            [
                ("account_type", "=", "asset_receivable"),
                ("company_ids", "in", cls.env.company.id),
                ("deprecated", "=", False),
            ],
            limit=1,
        )
        cls.sale_tax = cls.env["account.tax"].search(
            [
                ("company_id", "=", cls.env.company.id),
                ("type_tax_use", "=", "sale"),
                ("amount_type", "=", "percent"),
                ("amount", ">", 0),
                ("active", "=", True),
            ],
            limit=1,
        )
        cls.commercial_customer = cls.env["res.partner"].create(
            {
                "name": "Cliente comercial reporte",
                "is_company": True,
                "customer_rank": 1,
                "property_account_receivable_id": cls.receivable_account.id,
            }
        )
        cls.branch_customer = cls.env["res.partner"].create(
            {
                "name": "Sucursal facturada",
                "parent_id": cls.commercial_customer.id,
                "type": "invoice",
                "customer_rank": 1,
                "property_account_receivable_id": cls.receivable_account.id,
            }
        )
        cls.other_customer = cls.env["res.partner"].create(
            {
                "name": "Cliente fuera del reporte",
                "is_company": True,
                "customer_rank": 1,
                "property_account_receivable_id": cls.receivable_account.id,
            }
        )
        cls.product = cls.env["product.product"].search(
            [
                ("active", "=", True),
                ("uom_id", "=", cls.env.ref("uom.product_uom_unit").id),
            ],
            limit=1,
        )
        cls.service_product = cls.env["product.product"].search(
            [
                ("active", "=", True),
                ("type", "=", "service"),
                ("id", "!=", cls.product.id),
            ],
            limit=1,
        )
        if not cls.product or not cls.service_product:
            raise SkipTest(
                "Se requieren productos existentes para las pruebas del reporte."
            )
        cls.product.write(
            {
                "default_code": "RPT-001",
                "property_account_income_id": cls.income_account.id,
            }
        )
        cls.service_product.write(
            {
                "default_code": "RPT-SRV",
                "property_account_income_id": cls.income_account.id,
            }
        )
        cls.test_currency = cls.env["res.currency"].create(
            {
                "name": "XTS",
                "symbol": "XTS",
                "rounding": 0.01,
                "active": True,
            }
        )
        cls.env["res.currency.rate"].create(
            {
                "name": "2026-01-01",
                "rate": 2.0,
                "currency_id": cls.test_currency.id,
                "company_id": cls.env.company.id,
            }
        )

        cls.invoice = cls._create_document(
            "out_invoice",
            cls.branch_customer,
            cls.product,
            quantity=2.0,
            price_unit=100.0,
            discount=10.0,
            taxes=cls.sale_tax,
        )
        cls.credit_note = cls._create_document(
            "out_refund",
            cls.branch_customer,
            cls.product,
            quantity=1.0,
            price_unit=40.0,
            taxes=cls.sale_tax,
        )
        cls.dozen_invoice = cls._create_document(
            "out_invoice",
            cls.branch_customer,
            cls.product,
            quantity=1.0,
            price_unit=120.0,
            uom=cls.env.ref("uom.product_uom_dozen"),
            taxes=cls.env["account.tax"],
        )
        cls.currency_invoice = cls._create_document(
            "out_invoice",
            cls.branch_customer,
            cls.product,
            quantity=1.0,
            price_unit=50.0,
            currency=cls.test_currency,
            taxes=cls.env["account.tax"],
        )
        cls.service_invoice = cls._create_document(
            "out_invoice",
            cls.branch_customer,
            cls.service_product,
            quantity=1.0,
            price_unit=75.0,
            taxes=cls.env["account.tax"],
        )
        cls.unrelated_invoice = cls._create_document(
            "out_invoice",
            cls.other_customer,
            cls.product,
            quantity=99.0,
            price_unit=99.0,
            taxes=cls.env["account.tax"],
        )
        cls.draft_invoice = cls._create_document(
            "out_invoice",
            cls.branch_customer,
            cls.product,
            quantity=77.0,
            price_unit=77.0,
            taxes=cls.env["account.tax"],
            post=False,
        )

    @classmethod
    def _create_document(
        cls,
        move_type,
        partner,
        product,
        quantity,
        price_unit,
        discount=0.0,
        uom=None,
        currency=None,
        taxes=None,
        post=True,
    ):
        line_values = {
            "product_id": product.id,
            "name": product.display_name,
            "account_id": cls.income_account.id,
            "quantity": quantity,
            "price_unit": price_unit,
            "discount": discount,
            "tax_ids": [Command.set((taxes or cls.env["account.tax"]).ids)],
        }
        if uom:
            line_values["product_uom_id"] = uom.id
        move = cls.env["account.move"].create(
            {
                "move_type": move_type,
                "partner_id": partner.id,
                "invoice_date": "2026-01-15",
                "date": "2026-01-15",
                "currency_id": (
                    currency or cls.env.company.currency_id
                ).id,
                "invoice_line_ids": [Command.create(line_values)],
            }
        )
        if post:
            move.action_post()
        return move

    def _create_wizard(self, **values):
        values.setdefault("company_id", self.env.company.id)
        values.setdefault("partner_id", self.commercial_customer.id)
        values.setdefault("date_from", "2026-01-01")
        values.setdefault("date_to", "2026-01-31")
        return self.env["sng.customer.product.report.wizard"].create(values)

    def _create_product_customer_wizard(self, **values):
        values.setdefault("company_id", self.env.company.id)
        values.setdefault(
            "product_ids",
            [Command.set(self.product.ids)],
        )
        values.setdefault("date_from", "2026-01-01")
        values.setdefault("date_to", "2026-01-31")
        return self.env["sng.product.customer.report.wizard"].create(values)

    def test_snapshot_uses_posted_product_lines_and_commercial_customer(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()

        self.assertEqual(len(lines), 5)
        self.assertEqual(lines.partner_id, self.commercial_customer)
        self.assertEqual(lines.mapped("invoice_partner_id"), self.branch_customer)
        self.assertNotIn(self.unrelated_invoice, lines.mapped("move_id"))
        self.assertNotIn(self.draft_invoice, lines.mapped("move_id"))
        self.assertTrue(all(line.source_line_id.product_id for line in lines))
        self.assertTrue(
            all(line.source_line_id.display_type == "product" for line in lines)
        )
        self.assertIn(self.service_product, lines.mapped("product_id"))

    def test_invoice_values_match_accounting_line(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()
        report_line = lines.filtered(lambda line: line.move_id == self.invoice)
        source_line = self.invoice.invoice_line_ids.filtered(
            lambda line: line.display_type == "product"
        )

        self.assertEqual(report_line.quantity, source_line.quantity)
        self.assertEqual(report_line.price_unit, source_line.price_unit)
        self.assertEqual(report_line.discount, source_line.discount)
        self.assertEqual(report_line.subtotal, source_line.price_subtotal)
        self.assertAlmostEqual(
            report_line.tax_amount,
            source_line.price_total - source_line.price_subtotal,
            places=6,
        )
        self.assertEqual(report_line.total, source_line.price_total)

    def test_credit_note_is_negative_and_reduces_summary(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()
        credit_line = lines.filtered(
            lambda line: line.move_id == self.credit_note
        )

        self.assertTrue(credit_line.is_credit_note)
        self.assertLess(credit_line.quantity, 0)
        self.assertGreater(credit_line.price_unit, 0)
        self.assertLess(credit_line.subtotal, 0)
        self.assertLess(credit_line.tax_amount, 0)
        self.assertLess(credit_line.total, 0)

        product_summary = next(
            row
            for row in wizard._get_summary_rows()
            if row["product"] == self.product
        )
        self.assertEqual(product_summary["quantity"], 14.0)

    def test_currency_conversion_uses_document_date(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()
        report_line = lines.filtered(
            lambda line: line.move_id == self.currency_invoice
        )
        source_line = self.currency_invoice.invoice_line_ids.filtered(
            lambda line: line.display_type == "product"
        )
        expected = self.test_currency._convert(
            source_line.price_subtotal,
            self.env.company.currency_id,
            self.env.company,
            self.currency_invoice.invoice_date,
        )

        self.assertEqual(report_line.currency_id, self.test_currency)
        self.assertEqual(report_line.subtotal_company, expected)
        self.assertNotEqual(report_line.subtotal_company, report_line.subtotal)

    def test_product_filter_and_empty_result(self):
        wizard = self._create_wizard(
            product_ids=[Command.set(self.service_product.ids)]
        )
        lines = wizard._rebuild_lines()
        self.assertEqual(lines.product_id, self.service_product)

        empty_wizard = self._create_wizard(
            date_from="2025-01-01",
            date_to="2025-01-31",
        )
        with self.assertRaises(UserError):
            empty_wizard._rebuild_lines()

    def test_date_validation(self):
        with self.assertRaises(ValidationError):
            self._create_wizard(
                date_from="2026-02-01",
                date_to="2026-01-01",
            )

    def test_screen_action_uses_same_snapshot(self):
        wizard = self._create_wizard()
        action = wizard.action_view_report()

        self.assertEqual(
            action["res_model"], "sng.customer.product.report.line"
        )
        self.assertEqual(action["domain"], [("wizard_id", "=", wizard.id)])
        self.assertEqual(len(wizard.line_ids), 5)
        self.assertEqual(
            action["context"]["customer_product_report_wizard_id"],
            wizard.id,
        )

    def test_xlsx_has_summary_detail_and_negative_credit_note(self):
        wizard = self._create_wizard()
        wizard._rebuild_lines()
        report_data, report_type = self.env["ir.actions.report"]._render(
            "sng_customer_product_invoice_report.product_xlsx",
            wizard.ids,
            {},
        )
        workbook = load_workbook(BytesIO(report_data), read_only=True)

        self.assertEqual(report_type, "xlsx")
        self.assertEqual(
            workbook.sheetnames,
            ["Resumen por producto", "Detalle"],
        )
        summary_sheet = workbook["Resumen por producto"]
        detail_sheet = workbook["Detalle"]
        self.assertEqual(summary_sheet.cell(6, 1).value, "Código")
        self.assertEqual(detail_sheet.cell(6, 2).value, "Tipo")
        credit_rows = [
            row
            for row in detail_sheet.iter_rows(min_row=7, values_only=True)
            if row[1] == "Nota de crédito"
        ]
        self.assertEqual(len(credit_rows), 1)
        self.assertLess(credit_rows[0][7], 0)
        self.assertLess(credit_rows[0][11], 0)

    def test_pdf_template_renders_same_detail(self):
        wizard = self._create_wizard()
        wizard._rebuild_lines()
        html, report_type = self.env["ir.actions.report"]._render_qweb_html(
            "sng_customer_product_invoice_report."
            "report_customer_product_document",
            wizard.ids,
        )

        self.assertEqual(report_type, "html")
        self.assertIn(b"Productos facturados por cliente", html)
        self.assertIn(b"Nota de cr", html)
        self.assertIn(self.product.default_code.encode(), html)

    def test_readonly_accounting_user_can_generate_report(self):
        readonly_user = new_test_user(
            self.env,
            login="customer-product-report-readonly",
            groups="account.group_account_readonly",
            company_id=self.env.company.id,
            name="Customer Product Report Readonly",
        )
        readonly_env = self.env(user=readonly_user)
        wizard = readonly_env[
            "sng.customer.product.report.wizard"
        ].create(
            {
                "company_id": self.env.company.id,
                "partner_id": self.commercial_customer.id,
                "date_from": "2026-01-01",
                "date_to": "2026-01-31",
            }
        )

        lines = wizard._rebuild_lines()
        self.assertEqual(len(lines), 5)

    def test_company_access_and_result_rule(self):
        wizard = self._create_wizard()
        wizard._check_company_access()
        rule = self.env.ref(
            "sng_customer_product_invoice_report."
            "customer_product_report_line_company_rule"
        )
        self.assertIn("company_ids", rule.domain_force)

    def test_product_customer_report_finds_all_customers(self):
        wizard = self._create_product_customer_wizard()
        lines = wizard._rebuild_lines()

        self.assertEqual(len(lines), 5)
        self.assertEqual(
            lines.mapped("partner_id"),
            self.commercial_customer | self.other_customer,
        )
        self.assertIn(self.unrelated_invoice, lines.mapped("move_id"))
        self.assertNotIn(self.draft_invoice, lines.mapped("move_id"))
        credit_line = lines.filtered("is_credit_note")
        self.assertEqual(len(credit_line), 1)
        self.assertLess(credit_line.quantity, 0)
        self.assertLess(credit_line.total, 0)

    def test_product_customer_summary_is_product_and_customer(self):
        wizard = self._create_product_customer_wizard()
        wizard._rebuild_lines()
        summary = wizard._get_summary_rows()

        self.assertEqual(len(summary), 2)
        quantities = {
            row["partner"].id: row["quantity"]
            for row in summary
        }
        self.assertEqual(quantities[self.commercial_customer.id], 14.0)
        self.assertEqual(quantities[self.other_customer.id], 99.0)

    def test_product_customer_requires_products_and_screen_snapshot(self):
        with self.assertRaises(ValidationError):
            self._create_product_customer_wizard(product_ids=[Command.clear()])

        wizard = self._create_product_customer_wizard()
        action = wizard.action_view_report()
        self.assertEqual(
            action["res_model"],
            "sng.product.customer.report.line",
        )
        self.assertEqual(action["domain"], [("wizard_id", "=", wizard.id)])
        self.assertEqual(len(wizard.line_ids), 5)

    def test_product_customer_pdf_and_xlsx(self):
        wizard = self._create_product_customer_wizard()
        wizard._rebuild_lines()

        html, html_type = self.env["ir.actions.report"]._render_qweb_html(
            "sng_customer_product_invoice_report."
            "report_product_customer_document",
            wizard.ids,
        )
        xlsx, xlsx_type = self.env["ir.actions.report"]._render(
            "sng_customer_product_invoice_report.sales_xlsx",
            wizard.ids,
            {},
        )
        workbook = load_workbook(BytesIO(xlsx), read_only=True)

        self.assertEqual(html_type, "html")
        self.assertIn(b"Clientes por producto facturado", html)
        self.assertIn(self.other_customer.name.encode(), html)
        self.assertEqual(xlsx_type, "xlsx")
        self.assertEqual(
            workbook.sheetnames,
            ["Clientes por producto", "Detalle"],
        )
        self.assertEqual(
            workbook["Clientes por producto"].cell(6, 3).value,
            "Cliente comercial",
        )

    def test_product_customer_readonly_user_can_export(self):
        readonly_user = new_test_user(
            self.env,
            login="product-customer-report-readonly",
            groups="account.group_account_readonly",
            company_id=self.env.company.id,
            name="Product Customer Report Readonly",
        )
        readonly_env = self.env(user=readonly_user)
        wizard = readonly_env[
            "sng.product.customer.report.wizard"
        ].create(
            {
                "company_id": self.env.company.id,
                "product_ids": [Command.set(self.product.ids)],
                "date_from": "2026-01-01",
                "date_to": "2026-01-31",
            }
        )

        lines = wizard._rebuild_lines()
        self.assertEqual(len(lines), 5)
