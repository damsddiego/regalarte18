# -*- coding: utf-8 -*-

from io import BytesIO

from openpyxl import load_workbook

from odoo.tests.common import TransactionCase


class TestWarehouseAvailabilityReport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.warehouse = cls.env["stock.warehouse"].search(
            [("company_id", "=", cls.env.company.id)],
            limit=1,
        )
        cls.group = cls.env["sng.warehouse.group"].create(
            {
                "name": "Almacenes principales",
                "warehouse_ids": [(6, 0, cls.warehouse.ids)],
            }
        )
        cls.product = cls.env["product.product"].create(
            {
                "name": "Producto disponible",
                "default_code": "DISP-001",
                "is_storable": True,
            }
        )
        cls.quant = cls.env["stock.quant"].create(
            {
                "product_id": cls.product.id,
                "location_id": cls.warehouse.lot_stock_id.id,
                "quantity": 12.0,
                "reserved_quantity": 2.0,
            }
        )

    def _create_wizard(self, **values):
        values.setdefault("warehouse_group_id", self.group.id)
        return self.env["sng.warehouse.availability.wizard"].create(values)

    def test_group_defines_report_warehouses(self):
        wizard = self._create_wizard()
        self.assertEqual(wizard._get_warehouses(), self.warehouse)

    def test_report_uses_available_quantity(self):
        rows = self._create_wizard(product_ids=[(6, 0, self.product.ids)])._get_report_rows()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["quantity_by_warehouse"][self.warehouse.id], 10.0)

    def test_only_available_excludes_products_without_available_stock(self):
        self.quant.reserved_quantity = self.quant.quantity
        rows = self._create_wizard(
            product_ids=[(6, 0, self.product.ids)],
            only_available=True,
        )._get_report_rows()
        self.assertFalse(rows)

    def test_xlsx_renders_dynamic_warehouse_headers(self):
        wizard = self._create_wizard(product_ids=[(6, 0, self.product.ids)])
        report_data, report_type = self.env["ir.actions.report"]._render(
            "sng_warehouse_availability_report.availability_xlsx",
            wizard.ids,
            {},
        )
        sheet = load_workbook(BytesIO(report_data), read_only=True).active

        self.assertEqual(report_type, "xlsx")
        self.assertEqual(sheet.cell(1, 1).value, "Codigo")
        self.assertEqual(sheet.cell(2, 3).value, self.warehouse.name)
        self.assertEqual(sheet.cell(3, 1).value, self.product.default_code)

    def test_view_report_creates_filterable_lines(self):
        wizard = self._create_wizard(product_ids=[(6, 0, self.product.ids)])
        action = wizard.action_view_report()

        self.assertEqual(action["res_model"], "sng.warehouse.availability.line")
        self.assertEqual(action["domain"], [("wizard_id", "=", wizard.id)])
        self.assertEqual(len(wizard.line_ids), 1)
        self.assertEqual(wizard.line_ids.warehouse_id, self.warehouse)
        self.assertEqual(wizard.line_ids.product_id, self.product)
        self.assertEqual(wizard.line_ids.available_quantity, 10.0)
