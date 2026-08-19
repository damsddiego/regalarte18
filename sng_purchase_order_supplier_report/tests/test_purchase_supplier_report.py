# -*- coding: utf-8 -*-

from io import BytesIO

from openpyxl import load_workbook

from odoo import Command
from odoo.exceptions import UserError, ValidationError
from odoo.tests.common import TransactionCase, new_test_user


class TestPurchaseSupplierReport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.env = cls.env(context=dict(cls.env.context, tracking_disable=True))
        cls.company = cls.env.company
        cls.unit = cls.env.ref("uom.product_uom_unit")
        cls.dozen = cls.env.ref("uom.product_uom_dozen")
        cls.incoming_type = cls.env["stock.picking.type"].search(
            [
                ("code", "=", "incoming"),
                ("warehouse_id.company_id", "=", cls.company.id),
            ],
            limit=1,
        )
        cls.supplier_parent = cls.env["res.partner"].create(
            {
                "name": "Proveedor matriz reporte OC",
                "is_company": True,
                "supplier_rank": 1,
            }
        )
        cls.supplier_branch = cls.env["res.partner"].create(
            {
                "name": "Sucursal proveedor reporte OC",
                "parent_id": cls.supplier_parent.id,
                "supplier_rank": 1,
            }
        )
        cls.other_supplier = cls.env["res.partner"].create(
            {
                "name": "Otro proveedor reporte OC",
                "is_company": True,
                "supplier_rank": 1,
            }
        )
        cls.product = cls.env["product.product"].create(
            {
                "name": "Servicio reporte compra",
                "default_code": "PO-RPT-001",
                "type": "service",
                "purchase_ok": True,
                "uom_id": cls.unit.id,
                "uom_po_id": cls.unit.id,
            }
        )
        cls.product_2 = cls.env["product.product"].create(
            {
                "name": "Segundo servicio reporte compra",
                "default_code": "PO-RPT-002",
                "type": "service",
                "purchase_ok": True,
                "uom_id": cls.unit.id,
                "uom_po_id": cls.unit.id,
            }
        )
        cls.test_currency = cls.env["res.currency"].create(
            {
                "name": "XPR",
                "symbol": "XPR",
                "rounding": 0.01,
                "active": True,
            }
        )
        cls.env["res.currency.rate"].create(
            {
                "name": "2026-01-01",
                "rate": 2.0,
                "currency_id": cls.test_currency.id,
                "company_id": cls.company.id,
            }
        )

        cls.partial_order = cls._create_order(
            cls.supplier_branch,
            "2026-01-15 18:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 10.0,
                    "received": 4.0,
                    "price": 100.0,
                    "discount": 10.0,
                },
                {
                    "product": cls.product_2,
                    "qty": 5.0,
                    "received": 5.0,
                    "price": 20.0,
                },
            ],
            partner_ref="REF-PARCIAL",
        )
        cls.pending_currency_order = cls._create_order(
            cls.other_supplier,
            "2026-01-20 12:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 1.0,
                    "received": 0.0,
                    "price": 1200.0,
                    "uom": cls.dozen,
                }
            ],
            currency=cls.test_currency,
        )
        cls.over_received_order = cls._create_order(
            cls.other_supplier,
            "2026-01-22 12:00:00",
            lines=[
                {
                    "product": cls.product_2,
                    "qty": 5.0,
                    "received": 6.0,
                    "price": 30.0,
                }
            ],
        )
        cls.done_order = cls._create_order(
            cls.supplier_parent,
            "2026-01-25 12:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 2.0,
                    "received": 2.0,
                    "price": 50.0,
                }
            ],
            state="done",
        )
        cls.draft_order = cls._create_order(
            cls.supplier_parent,
            "2026-01-26 12:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 99.0,
                    "received": 0.0,
                    "price": 99.0,
                }
            ],
            state="draft",
        )
        cls.cancel_order = cls._create_order(
            cls.supplier_parent,
            "2026-01-27 12:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 88.0,
                    "received": 0.0,
                    "price": 88.0,
                }
            ],
            state="cancel",
        )
        cls.outside_order = cls._create_order(
            cls.supplier_parent,
            "2026-02-01 06:00:00",
            lines=[
                {
                    "product": cls.product,
                    "qty": 77.0,
                    "received": 0.0,
                    "price": 77.0,
                }
            ],
        )

    @classmethod
    def _create_order(
        cls,
        supplier,
        confirmation_date,
        lines,
        state="purchase",
        currency=None,
        partner_ref=None,
    ):
        order_values = {
            "partner_id": supplier.id,
            "company_id": cls.company.id,
            "currency_id": (currency or cls.company.currency_id).id,
            "date_order": confirmation_date,
            "date_approve": confirmation_date,
            "state": state,
            "partner_ref": partner_ref,
            "order_line": [],
        }
        if cls.incoming_type:
            order_values["picking_type_id"] = cls.incoming_type.id
        for index, values in enumerate(lines, start=1):
            uom = values.get("uom", values["product"].uom_po_id)
            order_values["order_line"].append(
                Command.create(
                    {
                        "name": values["product"].display_name,
                        "product_id": values["product"].id,
                        "product_qty": values["qty"],
                        "qty_received_manual": values["received"],
                        "product_uom": uom.id,
                        "price_unit": values["price"],
                        "discount": values.get("discount", 0.0),
                        "date_planned": "2026-03-%02d 12:00:00" % index,
                    }
                )
            )
        return cls.env["purchase.order"].create(order_values)

    def _create_wizard(self, **values):
        defaults = {
            "company_id": self.company.id,
            "date_from": "2026-01-01",
            "date_to": "2026-01-31",
            "reception_filter": "all",
        }
        defaults.update(values)
        return self.env["sng.purchase.supplier.report.wizard"].create(defaults)

    def test_all_filter_includes_only_confirmed_and_locked_orders(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()

        self.assertEqual(len(lines), 5)
        self.assertIn(self.partial_order, lines.mapped("order_id"))
        self.assertIn(self.done_order, lines.mapped("order_id"))
        self.assertNotIn(self.draft_order, lines.mapped("order_id"))
        self.assertNotIn(self.cancel_order, lines.mapped("order_id"))
        self.assertNotIn(self.outside_order, lines.mapped("order_id"))

    def test_transit_filter_keeps_only_pending_lines(self):
        wizard = self._create_wizard(reception_filter="transit")
        lines = wizard._rebuild_lines()

        self.assertEqual(len(lines), 2)
        partial_line = lines.filtered(
            lambda line: line.order_id == self.partial_order
        )
        self.assertEqual(len(partial_line), 1)
        self.assertEqual(partial_line.product_id, self.product)
        self.assertEqual(partial_line.qty_ordered, 10.0)
        self.assertEqual(partial_line.qty_received, 4.0)
        self.assertEqual(partial_line.qty_pending, 6.0)
        source_total = self.partial_order.order_line.filtered(
            lambda line: line.product_id == self.product
        ).price_total
        self.assertAlmostEqual(partial_line.total, source_total, places=2)
        self.assertAlmostEqual(
            partial_line.pending_total, source_total * 0.6, places=2
        )
        self.assertNotIn(self.over_received_order, lines.mapped("order_id"))

    def test_over_received_line_has_zero_pending_value(self):
        wizard = self._create_wizard()
        lines = wizard._rebuild_lines()
        line = lines.filtered(
            lambda report_line: report_line.order_id == self.over_received_order
        )

        self.assertEqual(line.qty_pending, 0.0)
        self.assertEqual(line.pending_subtotal, 0.0)
        self.assertEqual(line.pending_total_company, 0.0)
        self.assertEqual(line.reception_state, "received")

    def test_filter_uses_exact_supplier_contact(self):
        branch_wizard = self._create_wizard(
            supplier_ids=[Command.set(self.supplier_branch.ids)]
        )
        parent_wizard = self._create_wizard(
            supplier_ids=[Command.set(self.supplier_parent.ids)]
        )

        branch_lines = branch_wizard._rebuild_lines()
        parent_lines = parent_wizard._rebuild_lines()

        self.assertEqual(branch_lines.mapped("order_id"), self.partial_order)
        self.assertEqual(parent_lines.mapped("order_id"), self.done_order)

    def test_uom_and_currency_conversion_use_confirmation_date(self):
        wizard = self._create_wizard(
            supplier_ids=[Command.set(self.other_supplier.ids)],
            reception_filter="transit",
        )
        lines = wizard._rebuild_lines()
        line = lines.filtered(
            lambda report_line: report_line.order_id
            == self.pending_currency_order
        )

        self.assertEqual(line.base_qty_ordered, 12.0)
        expected_total = self.test_currency._convert(
            self.pending_currency_order.order_line.price_total,
            self.company.currency_id,
            self.company,
            line.confirmation_date.date(),
            round=False,
        )
        self.assertAlmostEqual(line.total_company, expected_total, places=2)
        self.assertAlmostEqual(
            line.pending_total_company, expected_total, places=2
        )

    def test_summary_groups_by_exact_supplier_and_product(self):
        wizard = self._create_wizard()
        wizard._rebuild_lines()
        summary = wizard._get_summary_rows()

        branch_product = next(
            row
            for row in summary
            if row["supplier"] == self.supplier_branch
            and row["product"] == self.product
        )
        self.assertEqual(branch_product["order_count"], 1)
        self.assertEqual(branch_product["qty_ordered"], 10.0)
        self.assertEqual(branch_product["qty_received"], 4.0)
        self.assertEqual(branch_product["qty_pending"], 6.0)

    def test_timezone_date_bounds_are_local_and_end_exclusive(self):
        wizard = self._create_wizard().with_context(tz="America/Costa_Rica")
        start, end = wizard._get_datetime_bounds()

        self.assertEqual(start.strftime("%Y-%m-%d %H:%M:%S"), "2026-01-01 06:00:00")
        self.assertEqual(end.strftime("%Y-%m-%d %H:%M:%S"), "2026-02-01 06:00:00")
        self.assertNotIn(self.outside_order.order_line, wizard._get_source_lines())

    def test_date_validation_and_empty_result(self):
        with self.assertRaises(ValidationError):
            self._create_wizard(
                date_from="2026-02-01", date_to="2026-01-01"
            )
        empty_wizard = self._create_wizard(
            date_from="2025-01-01", date_to="2025-01-31"
        )
        with self.assertRaises(UserError):
            empty_wizard._rebuild_lines()

    def test_screen_pdf_and_xlsx_outputs(self):
        wizard = self._create_wizard(reception_filter="transit")
        action = wizard.action_view_report()
        self.assertEqual(
            action["res_model"], "sng.purchase.supplier.report.line"
        )
        self.assertEqual(action["domain"], [("wizard_id", "=", wizard.id)])

        html, html_type = self.env["ir.actions.report"]._render_qweb_html(
            "sng_purchase_order_supplier_report."
            "report_purchase_supplier_document",
            wizard.ids,
        )
        xlsx, xlsx_type = self.env["ir.actions.report"]._render(
            "sng_po_supplier_report.xlsx", wizard.ids, {}
        )
        workbook = load_workbook(BytesIO(xlsx), read_only=True)

        self.assertEqual(html_type, "html")
        self.assertIn(b"Resumen de", html)
        self.assertIn(self.supplier_branch.name.encode(), html)
        self.assertEqual(xlsx_type, "xlsx")
        self.assertEqual(workbook.sheetnames, ["Resumen", "Detalle"])
        self.assertEqual(workbook["Resumen"].cell(8, 1).value, "Proveedor")
        self.assertEqual(workbook["Detalle"].cell(8, 4).value, "Orden de compra")

    def test_record_rules_isolate_user_snapshots(self):
        purchase_user = new_test_user(
            self.env,
            login="purchase-supplier-report-user",
            groups="purchase.group_purchase_user",
            company_id=self.company.id,
            name="Purchase Supplier Report User",
        )
        user_env = self.env(user=purchase_user)
        user_wizard = user_env["sng.purchase.supplier.report.wizard"].create(
            {
                "company_id": self.company.id,
                "date_from": "2026-01-01",
                "date_to": "2026-01-31",
                "reception_filter": "transit",
            }
        )
        user_lines = user_wizard._rebuild_lines()

        self.assertTrue(user_lines)
        self.assertEqual(user_lines.mapped("user_id"), purchase_user)
        observer_user = new_test_user(
            self.env,
            login="purchase-supplier-report-observer",
            groups="purchase.group_purchase_user",
            company_id=self.company.id,
            name="Purchase Supplier Report Observer",
        )
        observer_env = self.env(user=observer_user)
        self.assertNotIn(
            user_wizard,
            observer_env["sng.purchase.supplier.report.wizard"].search([]),
        )
        self.assertFalse(
            observer_env["sng.purchase.supplier.report.line"].search(
                [("wizard_id", "=", user_wizard.id)]
            )
        )

    def test_company_access_is_checked(self):
        wizard_rule = self.env.ref(
            "sng_purchase_order_supplier_report."
            "purchase_supplier_report_wizard_own_company_rule"
        )
        line_rule = self.env.ref(
            "sng_purchase_order_supplier_report."
            "purchase_supplier_report_line_own_company_rule"
        )
        for rule in wizard_rule | line_rule:
            self.assertIn("company_id", rule.domain_force)
            self.assertIn("company_ids", rule.domain_force)
            self.assertIn("user_id", rule.domain_force)
            self.assertIn("user.id", rule.domain_force)
