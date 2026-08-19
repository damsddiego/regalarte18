# -*- coding: utf-8 -*-

import base64
import unicodedata
from collections import defaultdict
from io import BytesIO

import openpyxl

from odoo import _, fields, models
from odoo.exceptions import UserError
from odoo.tools import float_compare


class SngStockScrapImportWizard(models.TransientModel):
    _name = "sng.stock.scrap.import.wizard"
    _description = "Importador de desecho de inventario"

    company_id = fields.Many2one(
        "res.company",
        string="Compania",
        required=True,
        default=lambda self: self.env.company,
    )
    file = fields.Binary(
        string="Archivo Excel",
        required=True,
        help="Sube un archivo .xlsx con las columnas Codigo, Descripcion y Danado.",
    )
    filename = fields.Char(string="Nombre del archivo")
    location_id = fields.Many2one(
        "stock.location",
        string="Ubicacion origen",
        required=True,
        default=lambda self: self._default_location_id(),
        domain="[('usage', '=', 'internal'), '|', ('company_id', '=', False), ('company_id', '=', company_id)]",
        help="Ubicacion interna desde donde se sacara el inventario danado.",
    )
    scrap_location_id = fields.Many2one(
        "stock.location",
        string="Ubicacion de desecho",
        required=True,
        default=lambda self: self._default_scrap_location_id(),
        domain="[('scrap_location', '=', True), '|', ('company_id', '=', False), ('company_id', '=', company_id)]",
        help="Ubicacion destino marcada como desecho en Odoo.",
    )
    origin = fields.Char(
        string="Origen",
        default=lambda self: _("Importacion desde Excel"),
        help="Texto que se guardara en el campo origen de cada desecho.",
    )
    validate_scraps = fields.Boolean(
        string="Validar automaticamente",
        default=True,
        help="Si esta activo, los desechos se validan al terminar la importacion.",
    )

    def _default_location_id(self):
        warehouse = self.env["stock.warehouse"].search(
            [("company_id", "=", self.env.company.id)],
            limit=1,
        )
        return warehouse.lot_stock_id.id

    def _default_scrap_location_id(self):
        return self.env["stock.location"].search(
            [
                ("scrap_location", "=", True),
                "|",
                ("company_id", "=", False),
                ("company_id", "=", self.env.company.id),
            ],
            limit=1,
        ).id

    @staticmethod
    def _normalize_header(value):
        text = unicodedata.normalize("NFKD", str(value or "").strip())
        return "".join(char for char in text if not unicodedata.combining(char)).lower()

    @staticmethod
    def _normalize_code(value):
        if value in (None, ""):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _format_error_list(errors, max_items=20):
        if len(errors) <= max_items:
            return "\n".join(errors)
        remaining = len(errors) - max_items
        return "\n".join(errors[:max_items] + [_("... y %s error(es) mas.") % remaining])

    def _load_sheet(self):
        self.ensure_one()
        try:
            workbook = openpyxl.load_workbook(BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception as error:
            raise UserError(_("No se pudo leer el archivo Excel: %s") % error) from error
        return workbook.active

    def _extract_rows(self, sheet):
        headers = [self._normalize_header(sheet.cell(row=1, column=column).value) for column in range(1, 4)]
        expected_headers = ["codigo", "descripcion", "danado"]
        if headers != expected_headers:
            raise UserError(
                _(
                    "Encabezados invalidos. Se esperaba: Codigo | Descripcion | Danado. "
                    "Encabezados detectados: %s"
                )
                % " | ".join(headers)
            )

        aggregated_rows = defaultdict(lambda: {"quantity": 0.0, "descriptions": set(), "rows": []})
        errors = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not row or not any(value not in (None, "") for value in row[:3]):
                continue

            code = self._normalize_code(row[0])
            description = str(row[1] or "").strip()
            quantity = row[2]

            if not code:
                errors.append(_("Fila %s: falta el codigo del producto.") % row_number)
                continue

            try:
                quantity = float(quantity or 0.0)
            except (TypeError, ValueError):
                errors.append(_("Fila %s: la cantidad danada no es numerica para %s.") % (row_number, code))
                continue

            if quantity <= 0:
                errors.append(_("Fila %s: la cantidad danada debe ser mayor que cero para %s.") % (row_number, code))
                continue

            aggregated_rows[code]["quantity"] += quantity
            if description:
                aggregated_rows[code]["descriptions"].add(description)
            aggregated_rows[code]["rows"].append(row_number)

        if errors:
            raise UserError(_("El archivo tiene errores:\n%s") % self._format_error_list(errors))

        if not aggregated_rows:
            raise UserError(_("El archivo no contiene lineas validas para importar."))

        return aggregated_rows

    def _get_products_by_code(self, codes):
        products = self.env["product.product"].search([("default_code", "in", list(codes))])
        products_by_code = defaultdict(lambda: self.env["product.product"])
        for product in products:
            products_by_code[product.default_code] |= product
        return products_by_code

    def _validate_rows(self, aggregated_rows):
        self.ensure_one()
        errors = []
        products_by_code = self._get_products_by_code(aggregated_rows.keys())
        prepared_rows = []

        for code, row_data in aggregated_rows.items():
            products = products_by_code.get(code, self.env["product.product"])
            if not products:
                errors.append(_("Codigo %s: no existe ningun producto con esa referencia interna.") % code)
                continue
            if len(products) > 1:
                errors.append(_("Codigo %s: existe mas de un producto con esa referencia interna.") % code)
                continue

            product = products[0]
            if product.tracking != "none":
                errors.append(
                    _("Codigo %s: el producto %s usa lotes/series y este archivo no incluye esa informacion.")
                    % (code, product.display_name)
                )
                continue

            if product.is_storable:
                available_qty = product.with_context(location=self.location_id.id, strict=True).qty_available
                if float_compare(
                    available_qty,
                    row_data["quantity"],
                    precision_rounding=product.uom_id.rounding,
                ) < 0:
                    errors.append(
                        _(
                            "Codigo %s: stock insuficiente en %s. Disponible: %s, solicitado: %s."
                        )
                        % (
                            code,
                            self.location_id.display_name,
                            available_qty,
                            row_data["quantity"],
                        )
                    )
                    continue

            prepared_rows.append(
                {
                    "code": code,
                    "product": product,
                    "quantity": row_data["quantity"],
                    "descriptions": sorted(row_data["descriptions"]),
                }
            )

        if errors:
            raise UserError(
                _(
                    "La importacion se detuvo porque hay datos pendientes por corregir:\n%s"
                )
                % self._format_error_list(errors)
            )

        return prepared_rows

    def action_import(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Debes subir un archivo Excel."))
        if not self.location_id:
            raise UserError(_("Debes seleccionar una ubicacion origen."))
        if not self.scrap_location_id:
            raise UserError(_("Debes seleccionar una ubicacion de desecho."))

        sheet = self._load_sheet()
        aggregated_rows = self._extract_rows(sheet)
        prepared_rows = self._validate_rows(aggregated_rows)

        origin = (self.origin or "").strip() or self.filename or _("Importacion desde Excel")
        scrap_values = []
        total_qty = 0.0
        for prepared_row in prepared_rows:
            product = prepared_row["product"]
            total_qty += prepared_row["quantity"]
            scrap_values.append(
                {
                    "company_id": self.company_id.id,
                    "product_id": product.id,
                    "product_uom_id": product.uom_id.id,
                    "scrap_qty": prepared_row["quantity"],
                    "location_id": self.location_id.id,
                    "scrap_location_id": self.scrap_location_id.id,
                    "origin": origin,
                }
            )

        scraps = self.env["stock.scrap"].create(scrap_values)
        if self.validate_scraps:
            for scrap in scraps:
                scrap.action_validate()

        action_name = _("Desechos importados") if self.validate_scraps else _("Desechos cargados")
        return {
            "type": "ir.actions.act_window",
            "name": action_name,
            "res_model": "stock.scrap",
            "view_mode": "list,form",
            "domain": [("id", "in", scraps.ids)],
            "context": {
                "default_location_id": self.location_id.id,
                "default_scrap_location_id": self.scrap_location_id.id,
                "default_origin": origin,
                "notification_info": {
                    "title": _("Importacion completada"),
                    "message": _(
                        "Se procesaron %s producto(s) por un total de %s unidad(es)."
                    )
                    % (len(scraps), total_qty),
                    "type": "success",
                },
            },
            "target": "current",
        }
