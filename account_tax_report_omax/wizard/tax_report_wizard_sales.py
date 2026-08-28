# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from datetime import datetime
import datetime as dt
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


class TaxReportWizardSales(models.TransientModel):
    _name = "tax.report.wizard.sales"
    _description = "Reporte de Impuestos de Ventas"

    start_date = fields.Date(string="Fecha inicial", required=True)
    end_date = fields.Date(string="Fecha final", default=datetime.today(), required=True)
    report_for = fields.Selection([
        ('partners', 'Contactos'),
        ('productCategory', 'Categoría de producto'),
        ('products', 'Productos'),
        ('sales_team', 'Equipo de ventas'),
        ('sales_person', 'Vendedor'),
        ('taxes', 'Impuestos')
    ], string='Reporte para')
    company_ids = fields.Many2many(
        comodel_name='res.company',
        relation='res_company_sales_rel',
        column1='wizard_id',
        column2='company_id',
        string="Compañía",
        required=True,
        default=lambda self: self.env.company,
        domain="[('id', 'in', context.get('allowed_company_ids'))]"
    )
    detailed_report = fields.Boolean(string='Reporte detallado')
    report_format = fields.Selection([
        ('excel', 'Excel'),
        ('pdf', 'PDF'),
    ], string='Formato del reporte', default='excel', required=True)

    partner_ids = fields.Many2many(
        comodel_name='res.partner',
        relation='res_partner_sales_rel',
        column1='wizard_id',
        column2='partner_id',
        string='Contactos'
    )
    product_category_ids = fields.Many2many(comodel_name='product.category', string='Categorías de producto')
    product_ids = fields.Many2many(comodel_name='product.product', string='Productos')
    sales_team_ids = fields.Many2many(comodel_name='crm.team', string='Equipos de ventas')
    salesperson_ids = fields.Many2many(comodel_name='res.users', string='Vendedores')
    tax_ids = fields.Many2many(
        comodel_name='account.tax',
        relation='account_tax_sales_rel',
        column1='wizard_id',
        column2='tax_id',
        string='Impuestos'
    )

    tax_ids_domain = fields.Text(string="Dominio de ventas", default=[])
    sales_person_ids_domain = fields.Text(string="Dominio de vendedores", default=[])
    generic_domain = fields.Text(string="Dominio general", default=[])
    partner_ids_domain = fields.Text(string="Dominio de contactos", default=[])

    @api.onchange('company_ids')
    def _onchange_companyIds_taxType(self):
        domain = []
        domain.append(('company_id.id', 'in', self.company_ids.ids))
        domain.append(('type_tax_use', '=', 'sale'))
        self.tax_ids_domain = domain
        self.tax_ids = [(5, 0, 0)]

    @api.onchange('company_ids')
    def _onchange_companyIds(self):
        domain = ['|', ('company_id', '=', False), ('company_id.id', 'in', self.company_ids.ids)]
        self.generic_domain = domain
        self.partner_ids = [(5, 0, 0)]
        self.product_ids = [(5, 0, 0)]
        self.sales_team_ids = [(5, 0, 0)]
        self.sales_person_ids_domain = [('company_ids', 'in', self.company_ids.ids)]
        self.partner_ids_domain = [('customer_rank', '>', 0)] + domain
        self.salesperson_ids = [(5, 0, 0)]

    def _get_base_domain(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ('move_type', 'in', ['out_invoice', 'out_refund']),
        ]
        return domain

    def _get_sales_breakdown_by_tarifa(self, invoices):
        """
        Calcula el desglose de IVA por tarifa considerando exoneraciones línea por línea.

        La exoneración se valida por:
          - partner.has_exoneration y vigencia
          - tax_code in ['01', '04'] (IVA general/bebidas)
          - CABYS del producto autorizado en allowed_cabys_ids del partner

        Retorna:
          tarifa_data: dict {rate_key: {rate, base, iva_bruto, exo, iva_neto}}
          invoice_data: list de dicts por factura para vista detallada
        """
        today = dt.date.today()
        tarifa_data = {}
        invoice_data = []

        for inv in invoices:
            in_out_refund = inv.move_type == 'out_refund'
            is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
            multiplier = -1 if (in_out_refund or is_rejected) else 1

            partner = inv.partner_id
            has_exo = bool(
                partner
                and getattr(partner, 'has_exoneration', False)
                and getattr(partner, 'percentage_exoneration', 0) > 0
            )
            if has_exo:
                exp = getattr(partner, 'date_expiration', None)
                if exp and exp < today:
                    has_exo = False

            allowed_codes = []
            exo_pct = 0.0
            if has_exo:
                allowed_codes = [c.name for c in getattr(partner, 'allowed_cabys_ids', [])]
                exo_pct = partner.percentage_exoneration

            inv_breakdown = {}

            product_lines = inv.invoice_line_ids.filtered(
                lambda l: l.display_type == 'product' and l.tax_ids
            )
            for line in product_lines:
                base = abs(line.price_subtotal)
                for tax in line.tax_ids:
                    tax_rate = tax.amount
                    rate_int = int(tax_rate) if tax_rate == int(tax_rate) else tax_rate
                    tax_key = f"{rate_int}%"

                    iva_bruto = base * (tax_rate / 100.0)
                    iva_exo = 0.0

                    if has_exo and getattr(tax, 'tax_code', '') in ['01', '04']:
                        cabys = None
                        if line.product_id:
                            cabys = getattr(line.product_id, 'cabys_code', None)
                            if not cabys and line.product_id.categ_id:
                                cabys = getattr(line.product_id.categ_id, 'cabys_code', None)

                        authorized = not allowed_codes or not cabys or cabys in allowed_codes
                        if authorized:
                            effective_exo = min(exo_pct, tax_rate)
                            iva_exo = base * (effective_exo / 100.0)

                    iva_neto = iva_bruto - iva_exo

                    base_s = base * multiplier
                    bruto_s = iva_bruto * multiplier
                    exo_s = iva_exo * multiplier
                    neto_s = iva_neto * multiplier

                    if tax_key not in tarifa_data:
                        tarifa_data[tax_key] = {'rate': tax_rate, 'base': 0.0, 'iva_bruto': 0.0, 'exo': 0.0, 'iva_neto': 0.0}
                    tarifa_data[tax_key]['base'] += base_s
                    tarifa_data[tax_key]['iva_bruto'] += bruto_s
                    tarifa_data[tax_key]['exo'] += exo_s
                    tarifa_data[tax_key]['iva_neto'] += neto_s

                    if tax_key not in inv_breakdown:
                        inv_breakdown[tax_key] = {'rate': tax_rate, 'base': 0.0, 'iva_bruto': 0.0, 'exo': 0.0, 'iva_neto': 0.0}
                    inv_breakdown[tax_key]['base'] += base_s
                    inv_breakdown[tax_key]['iva_bruto'] += bruto_s
                    inv_breakdown[tax_key]['exo'] += exo_s
                    inv_breakdown[tax_key]['iva_neto'] += neto_s

            if inv_breakdown:
                is_foreign = inv.currency_id != inv.company_id.currency_id
                if is_foreign:
                    # invoice_currency_rate almacena moneda_extranjera/CRC (ej: 0.001878 USD/CRC).
                    # Queremos mostrar CRC/moneda_extranjera (ej: 532.45), así que invertimos.
                    raw_rate = getattr(inv, 'invoice_currency_rate', None) or 0.0
                    if raw_rate:
                        rate = round(1.0 / raw_rate, 2)
                    else:
                        # Fallback desde montos: CRC / USD
                        total_usd = abs(getattr(inv, 'amount_total_in_currency_signed', 0.0) or 0.0)
                        total_crc = abs(inv.amount_total_signed or 0.0)
                        rate = round(total_crc / total_usd, 2) if total_usd else 0.0
                else:
                    rate = 1.0

                invoice_data.append({
                    'partner_name': inv.partner_id.name or 'Sin nombre',
                    'ref': inv.name,
                    'date': inv.invoice_date.strftime('%d/%m/%y'),
                    'currency': inv.currency_id.name,
                    'is_usd': inv.currency_id.name == 'USD',
                    'is_foreign': is_foreign,
                    'currency_rate': rate,
                    'state_tributacion': getattr(inv, 'state_tributacion', '') or '',
                    'has_exoneration': getattr(inv, 'has_exoneration', False),
                    'breakdown': inv_breakdown,
                })

        tarifa_sorted = dict(sorted(tarifa_data.items(), key=lambda x: x[1]['rate']))
        return tarifa_sorted, invoice_data

    def _write_cell(self, worksheet, row, col, value, style_name):
        cell = worksheet.cell(row=row, column=col)
        cell.value = value

        if 'header' in style_name:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        elif 'bold' in style_name:
            cell.font = Font(bold=True)

        if 'green' in style_name:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        if 'yellow' in style_name:
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        if 'center' in style_name:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif 'right' in style_name:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif 'left' in style_name:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    def print_xlsx_report(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'IVA Ventas por Tarifa'

        row = 1
        col = 1

        self._write_cell(worksheet, row, col, 'Reporte de IVA por Tarifa — Ventas', 'header_center_bold')
        row += 2
        self._write_cell(worksheet, row, col, f'Desde: {self.start_date}   Hasta: {self.end_date}', 'left_bold')
        row += 2

        if self.report_for == 'taxes':
            domain = self._get_base_domain()
            if self.partner_ids:
                domain.append(('partner_id.id', 'in', self.partner_ids.ids))
            if self.tax_ids:
                domain.append(('invoice_line_ids.tax_ids', 'in', self.tax_ids.ids))

            invoices = self.env['account.move'].sudo().search(domain, order='invoice_date, name')
            if not invoices:
                raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

            tarifa_data, invoice_data = self._get_sales_breakdown_by_tarifa(invoices)
            all_rates = list(tarifa_data.keys())

            if not self.detailed_report:
                # ── RESUMEN POR TARIFA ──────────────────────────────────────
                headers = [
                    (_('Tarifa'), 'header_center_bold'),
                    (_('Base Gravable'), 'header_right'),
                    (_('IVA sin Exoneración'), 'header_right'),
                    (_('Monto Exonerado'), 'header_right'),
                    (_('IVA Cobrado'), 'header_right'),
                ]
                for i, (h, s) in enumerate(headers):
                    self._write_cell(worksheet, row, col + i, h, s)
                row += 1

                grand = {'base': 0.0, 'iva_bruto': 0.0, 'exo': 0.0, 'iva_neto': 0.0}
                for rate_key, data in tarifa_data.items():
                    self._write_cell(worksheet, row, col, rate_key, 'center_bold')
                    self._write_cell(worksheet, row, col + 1, round(data['base'], 2), 'right')
                    self._write_cell(worksheet, row, col + 2, round(data['iva_bruto'], 2), 'right')
                    self._write_cell(worksheet, row, col + 3, round(data['exo'], 2), 'right')
                    self._write_cell(worksheet, row, col + 4, round(data['iva_neto'], 2), 'right')
                    row += 1
                    for k in grand:
                        grand[k] += data[k]

                row += 1
                self._write_cell(worksheet, row, col, 'TOTAL', 'header_center_bold')
                self._write_cell(worksheet, row, col + 1, round(grand['base'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 2, round(grand['iva_bruto'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 3, round(grand['exo'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 4, round(grand['iva_neto'], 2), 'header_right')

            else:
                # ── DETALLADO: MATRIZ POR FACTURA ───────────────────────────
                # Encabezados fijos
                fixed_headers = [
                    (_('Fecha'), 'header_center_bold'),
                    (_('Cliente'), 'header_center_bold'),
                    (_('Factura'), 'header_center_bold'),
                    (_('Exo'), 'header_center_bold'),
                    (_('Moneda'), 'header_center_bold'),
                    (_('T.C.'), 'header_right'),
                ]
                for i, (h, s) in enumerate(fixed_headers):
                    self._write_cell(worksheet, row, col + i, h, s)
                current_col = col + len(fixed_headers)

                # Encabezados dinámicos por tarifa (Base + IVA Cobrado)
                rate_col_map = {}
                for rate_key in all_rates:
                    rate_col_map[rate_key] = current_col
                    self._write_cell(worksheet, row, current_col, f'Base {rate_key}', 'header_right')
                    self._write_cell(worksheet, row, current_col + 1, f'IVA {rate_key}', 'header_right')
                    current_col += 2

                # Columnas de totales
                col_total_base = current_col
                col_total_exo = current_col + 1
                col_total_iva = current_col + 2
                col_estado = current_col + 3
                self._write_cell(worksheet, row, col_total_base, _('Total Base'), 'header_right')
                self._write_cell(worksheet, row, col_total_exo, _('Total Exonerado'), 'header_right')
                self._write_cell(worksheet, row, col_total_iva, _('IVA Cobrado Total'), 'header_right')
                self._write_cell(worksheet, row, col_estado, _('Estado Trib.'), 'header_center_bold')
                row += 1

                # Acumuladores de columnas
                grand_base = 0.0
                grand_exo = 0.0
                grand_iva = 0.0
                grand_by_rate = {rk: {'base': 0.0, 'iva_neto': 0.0} for rk in all_rates}

                for inv_data in invoice_data:
                    style_prefix = 'green_' if inv_data['is_usd'] else ''
                    exo_mark = '✓' if inv_data['has_exoneration'] else ''
                    breakdown = inv_data['breakdown']

                    row_base = 0.0
                    row_exo = 0.0
                    row_iva = 0.0

                    self._write_cell(worksheet, row, col, inv_data['date'], style_prefix + 'center')
                    self._write_cell(worksheet, row, col + 1, inv_data['partner_name'], style_prefix + 'left')
                    self._write_cell(worksheet, row, col + 2, inv_data['ref'], style_prefix + 'left')
                    self._write_cell(worksheet, row, col + 3, exo_mark, style_prefix + 'center')
                    self._write_cell(worksheet, row, col + 4, inv_data['currency'], style_prefix + 'center')
                    tc_value = round(inv_data['currency_rate'], 5) if inv_data['is_foreign'] else ''
                    self._write_cell(worksheet, row, col + 5, tc_value, style_prefix + 'right')

                    for rate_key in all_rates:
                        c = rate_col_map[rate_key]
                        if rate_key in breakdown:
                            b = round(breakdown[rate_key]['base'], 2)
                            n = round(breakdown[rate_key]['iva_neto'], 2)
                            e = round(breakdown[rate_key]['exo'], 2)
                            self._write_cell(worksheet, row, c, b, style_prefix + 'right')
                            self._write_cell(worksheet, row, c + 1, n, style_prefix + 'right')
                            row_base += b
                            row_exo += e
                            row_iva += n
                            grand_by_rate[rate_key]['base'] += b
                            grand_by_rate[rate_key]['iva_neto'] += n
                        else:
                            self._write_cell(worksheet, row, c, '-', style_prefix + 'center')
                            self._write_cell(worksheet, row, c + 1, '-', style_prefix + 'center')

                    self._write_cell(worksheet, row, col_total_base, round(row_base, 2), style_prefix + 'right')
                    self._write_cell(worksheet, row, col_total_exo, round(row_exo, 2), style_prefix + 'right')
                    self._write_cell(worksheet, row, col_total_iva, round(row_iva, 2), style_prefix + 'right')
                    self._write_cell(worksheet, row, col_estado, inv_data['state_tributacion'], style_prefix + 'left')

                    grand_base += row_base
                    grand_exo += row_exo
                    grand_iva += row_iva
                    row += 1

                # Fila de totales generales
                row += 1
                self._write_cell(worksheet, row, col + 1, 'TOTAL GENERAL', 'header_center_bold')
                for rate_key in all_rates:
                    c = rate_col_map[rate_key]
                    self._write_cell(worksheet, row, c, round(grand_by_rate[rate_key]['base'], 2), 'header_right')
                    self._write_cell(worksheet, row, c + 1, round(grand_by_rate[rate_key]['iva_neto'], 2), 'header_right')
                self._write_cell(worksheet, row, col_total_base, round(grand_base, 2), 'header_right')
                self._write_cell(worksheet, row, col_total_exo, round(grand_exo, 2), 'header_right')
                self._write_cell(worksheet, row, col_total_iva, round(grand_iva, 2), 'header_right')

                # Resumen por tarifa al pie
                row += 3
                self._write_cell(worksheet, row, col, 'Resumen por Tarifa', 'header_center_bold')
                row += 1
                summary_headers = [
                    _('Tarifa'), _('Base Gravable'), _('IVA sin Exoneración'),
                    _('Monto Exonerado'), _('IVA Cobrado'),
                ]
                for i, h in enumerate(summary_headers):
                    self._write_cell(worksheet, row, col + i, h, 'header_right' if i > 0 else 'header_center_bold')
                row += 1
                grand2 = {'base': 0.0, 'iva_bruto': 0.0, 'exo': 0.0, 'iva_neto': 0.0}
                for rate_key, data in tarifa_data.items():
                    self._write_cell(worksheet, row, col, rate_key, 'center_bold')
                    self._write_cell(worksheet, row, col + 1, round(data['base'], 2), 'right')
                    self._write_cell(worksheet, row, col + 2, round(data['iva_bruto'], 2), 'right')
                    self._write_cell(worksheet, row, col + 3, round(data['exo'], 2), 'right')
                    self._write_cell(worksheet, row, col + 4, round(data['iva_neto'], 2), 'right')
                    row += 1
                    for k in grand2:
                        grand2[k] += data[k]
                self._write_cell(worksheet, row, col, 'TOTAL', 'header_center_bold')
                self._write_cell(worksheet, row, col + 1, round(grand2['base'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 2, round(grand2['iva_bruto'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 3, round(grand2['exo'], 2), 'header_right')
                self._write_cell(worksheet, row, col + 4, round(grand2['iva_neto'], 2), 'header_right')

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)

        export_id = self.env['account.excel.tax.report'].create({
            'excel_file': base64.b64encode(fp.getvalue()),
            'file_name': 'iva_ventas_por_tarifa.xlsx'
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'account.excel.tax.report',
            'view_mode': 'form',
            'res_id': export_id.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

    def print_report(self):
        self.ensure_one()
        if self.report_format == 'excel':
            return self.print_xlsx_report()
        elif self.report_format == 'pdf':
            raise models.UserError(_("PDF report not implemented yet"))
