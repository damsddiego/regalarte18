# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from time import gmtime, strftime
from datetime import datetime
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from json import dumps
import ast
import json
from io import BytesIO



class AccountExcelTaxRepost(models.TransientModel):
    _name = "account.excel.tax.report"
    _description = 'Reporte de impuestos en Excel'

    excel_file = fields.Binary('Reporte Excel')
    file_name = fields.Char('Nombre del archivo del reporte',size=64,readonly=True)



class tax_report_wizard_account(models.TransientModel):
    _name = "tax.report.wizard.account"
    _description = "Reporte de Impuestos Contables"

    start_date = fields.Date(string="Fecha inicial", required=True)
    end_date = fields.Date(string="Fecha final", default=datetime.today(), required=True)
    tax_type = fields.Selection([('sales', 'Ventas'),('purchases', 'Compras')], string='Tipo de impuesto',)
    report_for = fields.Selection([('partners', 'Contactos'),('productCategory', 'Categoría de producto'),('products', 'Productos'),('sales_team', 'Equipo de ventas'),('sales_person', 'Vendedor'),('taxes', 'Impuestos')], string='Reporte para',)
    company_ids = fields.Many2many(
            comodel_name='res.company',relation='res_company_rel',column1='wizard_id',column2='company_id',string="Compañía",required=True,default=lambda self: self.env.company,domain="[('id', 'in', context.get('allowed_company_ids'))]")
    detailed_report = fields.Boolean(string='Reporte detallado',)
    report_format = fields.Selection([('excel', 'Excel'),('pdf', 'PDF'),], string='Formato del reporte', default='excel', required=True)

    partner_ids = fields.Many2many(comodel_name='res.partner',relation='res_partner_rel',column1='wizard_id',column2='partner_id', string='Contactos',)

    product_category_ids = fields.Many2many(comodel_name='product.category', string='Categorías de producto')
    product_ids = fields.Many2many(comodel_name='product.product', string='Productos')
    sales_team_ids = fields.Many2many(comodel_name='crm.team', string='Equipos de ventas')
    salesperson_ids = fields.Many2many(comodel_name='res.users', string='Vendedores')
    tax_ids = fields.Many2many(comodel_name='account.tax', string='Impuestos')

    tax_ids_domain = fields.Text(string="Dominio de ventas/compras",default=[])
    sales_person_ids_domain = fields.Text(string="Dominio de vendedores",default=[])
    generic_domain = fields.Text(string="Dominio general",default=[])
    partner_ids_domain = fields.Text(string="Dominio de contactos",default=[],)

    @api.onchange('company_ids','tax_type')
    def _onchange_companyIds_taxType(self):
        domain = []
        domain.append(('company_id.id', 'in', self.company_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('type_tax_use', '=', 'sale'))
        elif self.tax_type == 'purchases':
            domain.append(('type_tax_use', '=', 'purchase'))
        self.tax_ids_domain = domain
        self.tax_ids = [(5, 0, 0)]

    @api.onchange('company_ids')
    def _onchange_companyIds(self):
        domain = ['|', ('company_id', '=', False), ('company_id.id', 'in', self.company_ids.ids)]
        self.generic_domain = domain
        self.partner_ids = [(5, 0, 0)]
        self.product_ids = [(5, 0, 0)]
        self.sales_team_ids = [(5, 0, 0)]
        self.sales_person_ids_domain = [('company_ids', 'in', self.company_ids.ids)]
        self.partner_ids_domain = ['|',('supplier_rank', '>', 0),('customer_rank', '>', 0)] + domain
        self.salesperson_ids = [(5, 0, 0)]


    def _filter_inv_on_tax(self,tax,invoice):
        line = []
        for inv in invoice:
            in_out_refund = inv.move_type in ['out_refund','in_refund']
            is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
            is_negative = in_out_refund or is_rejected
            taxes = inv.tax_totals
            if taxes['subtotals'][0]['tax_groups']:
                for tax_group in taxes['subtotals'][0]['tax_groups']:
                    if tax_group.get('group_name') == tax:
                        group = tax_group

                        # Get all amount fields from invoice with safe access
                        amount_discount_electronic_invoice = getattr(inv, 'amount_discount_electronic_invoice', 0.0) or 0.0
                        amount_iva_returned = getattr(inv, 'amount_iva_returned', 0.0) or 0.0
                        amount_paid = getattr(inv, 'amount_paid', 0.0) or 0.0
                        amount_residual = getattr(inv, 'amount_residual', 0.0) or 0.0
                        amount_residual_signed = getattr(inv, 'amount_residual_signed', 0.0) or 0.0
                        amount_subtotal_without_iva_returned = getattr(inv, 'amount_subtotal_without_iva_returned', 0.0) or 0.0
                        amount_tax = getattr(inv, 'amount_tax', 0.0) or 0.0
                        amount_tax_electronic_invoice = getattr(inv, 'amount_tax_electronic_invoice', 0.0) or 0.0
                        amount_tax_signed = getattr(inv, 'amount_tax_signed', 0.0) or 0.0
                        amount_total = getattr(inv, 'amount_total', 0.0) or 0.0
                        amount_total_in_currency_signed = getattr(inv, 'amount_total_in_currency_signed', 0.0) or 0.0
                        amount_total_signed = getattr(inv, 'amount_total_signed', 0.0) or 0.0
                        amount_untaxed = getattr(inv, 'amount_untaxed', 0.0) or 0.0
                        amount_untaxed_in_currency_signed = getattr(inv, 'amount_untaxed_in_currency_signed', 0.0) or 0.0
                        state_tributacion = getattr(inv, 'state_tributacion', '') or ''

                        # Apply refund logic to amounts
                        discount_adjusted = amount_discount_electronic_invoice if not is_negative else -abs(amount_discount_electronic_invoice)
                        total_signed_adjusted = amount_total_signed if not is_negative else -abs(amount_total_signed)

                        # Calculate Subtotal (Total firmado + Descuento factura electrónica)
                        sub_total = total_signed_adjusted + discount_adjusted

                        val = {
                            'name' : inv.partner_id.name,
                            'ref' : inv.name,
                            'inv_date' : inv.invoice_date.strftime("%d/%m/%y"),
                            'sub_total' : sub_total,
                            'amount_discount_electronic_invoice' : discount_adjusted,
                            'amount_iva_returned' : amount_iva_returned if not is_negative else -abs(amount_iva_returned),
                            'amount_paid' : amount_paid if not is_negative else -abs(amount_paid),
                            'amount_residual' : amount_residual if not is_negative else -abs(amount_residual),
                            'amount_residual_signed' : amount_residual_signed if not is_negative else -abs(amount_residual_signed),
                            'amount_subtotal_without_iva_returned' : amount_subtotal_without_iva_returned if not is_negative else -abs(amount_subtotal_without_iva_returned),
                            'amount_tax' : amount_tax if not is_negative else -abs(amount_tax),
                            'amount_tax_electronic_invoice' : amount_tax_electronic_invoice if not is_negative else -abs(amount_tax_electronic_invoice),
                            'amount_tax_signed' : amount_tax_signed if not is_negative else -abs(amount_tax_signed),
                            'amount_total' : amount_total if not is_negative else -abs(amount_total),
                            'amount_total_in_currency_signed' : amount_total_in_currency_signed if not is_negative else -abs(amount_total_in_currency_signed),
                            'amount_total_signed' : total_signed_adjusted,
                            'amount_untaxed' : amount_untaxed if not is_negative else -abs(amount_untaxed),
                            'amount_untaxed_in_currency_signed' : amount_untaxed_in_currency_signed if not is_negative else -abs(amount_untaxed_in_currency_signed),
                            'state_tributacion' : state_tributacion,
                        }
                        line.append(val)
        return line

    def get_report_datas_partners(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.partner_ids:
            domain.append(('partner_id.id', 'in', self.partner_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        partner_list = invoice_lines.mapped('partner_id').ids

        partners = self.env['res.partner'].sudo().search([('id', 'in', partner_list)])

        data_dics = []
        tex_gr_names = []
        for p in partners:
            invoice = invoice_lines.filtered(lambda a: a.partner_id.id == p.id)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            partner_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            partner_dic['name'] = p.name if p.name else 'Nombre no disponible'
            partner_dic['bal_total_untax'] = bal_total_untax
            partner_dic['total_tax_paid'] = total_tax_paid
            partner_dic['total_tax_receievd'] = total_tax_receievd
            partner_dic['balance_total_tax'] = balance_total_tax
            partner_dic['tax_groups'] = tax_groups
            data_dics.append(partner_dic)
        return data_dics

    def get_report_datas_productCategories(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.product_category_ids:
            domain.append(('invoice_line_ids.product_id.categ_id', 'in', self.product_category_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        categ_lst = invoice_lines.mapped('invoice_line_ids').mapped('product_id').mapped('categ_id').ids

        categ_ids = self.env['product.category'].sudo().search([('id', 'in', categ_lst)])

        data_dics = []
        tex_gr_names = []
        for categ in categ_ids:
            prods_ids = self.env['product.product'].sudo().search([('categ_id', '=', categ.id)])
            invoice = invoice_lines.filtered(lambda a: a.invoice_line_ids.product_id.id in prods_ids.ids)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            categ_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            categ_dic['name'] = categ.name if categ.name else 'Nombre no disponible'
            categ_dic['bal_total_untax'] = bal_total_untax
            categ_dic['total_tax_paid'] = total_tax_paid
            categ_dic['total_tax_receievd'] = total_tax_receievd
            categ_dic['balance_total_tax'] = balance_total_tax
            categ_dic['tax_groups'] = tax_groups
            data_dics.append(categ_dic)
        return data_dics

    def get_report_datas_products(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.product_ids:
            domain.append(('invoice_line_ids.product_id', 'in', self.product_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        prod_lst = invoice_lines.mapped('invoice_line_ids').mapped('product_id').ids

        prod_ids = self.env['product.product'].sudo().search([('id', 'in', prod_lst)])

        data_dics = []
        tex_gr_names = []
        for prod in prod_ids:
            invoice = invoice_lines.filtered(lambda a: a.invoice_line_ids.product_id.id == prod.id)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            prod_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            prod_dic['name'] = prod.name if prod.name else 'Nombre no disponible'
            prod_dic['bal_total_untax'] = bal_total_untax
            prod_dic['total_tax_paid'] = total_tax_paid
            prod_dic['total_tax_receievd'] = total_tax_receievd
            prod_dic['balance_total_tax'] = balance_total_tax
            prod_dic['tax_groups'] = tax_groups
            data_dics.append(prod_dic)
        return data_dics

    def get_report_datas_sales_team(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.sales_team_ids:
            domain.append(('team_id', 'in', self.sales_team_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        team_lst = invoice_lines.mapped('team_id').ids

        team_ids = self.env['crm.team'].sudo().search([('id', 'in', team_lst)])

        data_dics = []
        tex_gr_names = []
        for team in team_ids:
            invoice = invoice_lines.filtered(lambda a: a.team_id.id == team.id)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            team_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            team_dic['name'] = team.name if team.name else 'Nombre no disponible'
            team_dic['bal_total_untax'] = bal_total_untax
            team_dic['total_tax_paid'] = total_tax_paid
            team_dic['total_tax_receievd'] = total_tax_receievd
            team_dic['balance_total_tax'] = balance_total_tax
            team_dic['tax_groups'] = tax_groups
            data_dics.append(team_dic)
        return data_dics

    def get_report_datas_sales_person(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.salesperson_ids:
            domain.append(('invoice_user_id', 'in', self.salesperson_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        sales_person_lst = invoice_lines.mapped('invoice_user_id').ids

        sales_person_ids = self.env['res.users'].sudo().search([('id', 'in', sales_person_lst)])

        data_dics = []
        tex_gr_names = []
        for sales_person in sales_person_ids:
            invoice = invoice_lines.filtered(lambda a: a.invoice_user_id.id == sales_person.id)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            sales_person_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            sales_person_dic['name'] = sales_person.name if sales_person.name else 'Nombre no disponible'
            sales_person_dic['bal_total_untax'] = bal_total_untax
            sales_person_dic['total_tax_paid'] = total_tax_paid
            sales_person_dic['total_tax_receievd'] = total_tax_receievd
            sales_person_dic['balance_total_tax'] = balance_total_tax
            sales_person_dic['tax_groups'] = tax_groups
            data_dics.append(sales_person_dic)
        return data_dics

    def get_report_datas_taxes(self):
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ]
        if self.tax_ids:
            domain.append(('invoice_line_ids.tax_ids', 'in', self.tax_ids.ids))
        if self.tax_type == 'sales':
            domain.append(('move_type', 'in', ['out_invoice', 'out_refund']))
        elif self.tax_type == 'purchases':
            domain.append(('move_type', 'in', ['in_invoice', 'in_refund']))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("No hay facturas para la selección indicada. Revise los filtros."))

        tax_lst = invoice_lines.mapped('invoice_line_ids').mapped('tax_ids').ids

        tax_ids = self.env['account.tax'].sudo().search([('id', 'in', tax_lst)])

        data_dics = []
        tex_gr_names = []
        for tax in tax_ids:
            invoice = invoice_lines.filtered(lambda a: tax in a.invoice_line_ids.tax_ids)
            bal_total_untax = 0
            total_tax_paid = 0
            total_tax_receievd = 0
            balance_total_tax = 0
            tax_groups = []
            tax_dic = {}
            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund','in_refund']
                is_rejected = getattr(inv, 'state_tributacion', '') == 'rechazado'
                is_negative = in_out_refund or is_rejected
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not is_negative else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            #pdb.set_trace()
                            if inv.move_type in ['in_invoice', 'in_refund']:
                                balance_total_tax -= tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_paid += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                            elif inv.move_type in ['out_invoice', 'out_refund']:
                                balance_total_tax += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))
                                total_tax_receievd += tax_group.get('tax_amount') if not is_negative else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name,invoice),
                    })
            tax_dic['name'] = tax.name if tax.name else 'Nombre no disponible'
            tax_dic['bal_total_untax'] = bal_total_untax
            tax_dic['total_tax_paid'] = total_tax_paid
            tax_dic['total_tax_receievd'] = total_tax_receievd
            tax_dic['balance_total_tax'] = balance_total_tax
            tax_dic['tax_groups'] = tax_groups
            data_dics.append(tax_dic)
        return data_dics

    def _write_cell(self, worksheet, row, col, value, style_type='normal'):
        """Helper method to write cells with proper formatting for openpyxl"""
        cell = worksheet.cell(row=row+1, column=col+1)

        # Write value - if it's a numeric value, write as number
        if isinstance(value, (int, float)):
            cell.value = value
            cell.number_format = '#,##0.00'
        else:
            cell.value = value

        # Apply styles
        if style_type == 'header_bold':
            cell.font = Font(bold=True, size=15)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif style_type == 'header_right':
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif style_type == 'header_center_bold':
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif style_type == 'left_bold':
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
        elif style_type == 'right_bold':
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        elif style_type == 'left':
            cell.alignment = Alignment(horizontal='left')
        elif style_type == 'right':
            cell.alignment = Alignment(horizontal='right')
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        elif style_type == 'center_bold':
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        return cell

    def print_xlsx_report(self):
        for company in self.company_ids:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Reporte de Impuestos'

            # STYLES - No longer needed with openpyxl, using _write_cell helper

            row = 0
            col = 0

            # Merge cells for header
            worksheet.merge_cells(start_row=row+1, start_column=col+1, end_row=row+1, end_column=col+5)
            self._write_cell(worksheet, row, col, _("Reporte de Impuestos"), 'header_bold')
            row+=2

            self._write_cell(worksheet, row, col, _("Compañía:"), 'left_bold')
            self._write_cell(worksheet, row, col+1, _(company.name), 'left')
            row+=1

            self._write_cell(worksheet, row, col, _("Fecha inicial:"), 'left_bold')
            self._write_cell(worksheet, row, col+1, str(self.start_date), 'left')
            row+=1

            self._write_cell(worksheet, row, col, _("Fecha final:"), 'left_bold')
            self._write_cell(worksheet, row, col+1, str(self.end_date), 'left')
            row+=2

            if self.report_for == 'partners':

                self._write_cell(worksheet, row, col, 'Contacto', 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_partners()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, 'Cliente', 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, 'Factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, 'Fecha de factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, 'Subtotal', 'header_right')
                            self._write_cell(worksheet, row, col+6, 'Descuento factura electrónica', 'header_right')
                            self._write_cell(worksheet, row, col+7, 'Subtotal sin IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+8, 'IVA', 'header_right')
                            self._write_cell(worksheet, row, col+9, 'IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+10, 'Total firmado', 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, 'Estado Tributación', 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount')
                                total_tax += line.get('tax_amount')
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            elif self.report_for == 'productCategory':

                self._write_cell(worksheet, row, col, 'Categoría de producto', 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_productCategories()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, 'Cliente', 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, 'Factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, 'Fecha de factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, 'Subtotal', 'header_right')
                            self._write_cell(worksheet, row, col+6, 'Descuento factura electrónica', 'header_right')
                            self._write_cell(worksheet, row, col+7, 'Subtotal sin IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+8, 'IVA', 'header_right')
                            self._write_cell(worksheet, row, col+9, 'IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+10, 'Total firmado', 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, 'Estado Tributación', 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount')
                                total_tax += line.get('tax_amount')
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            elif self.report_for == 'products':

                self._write_cell(worksheet, row, col, 'Producto', 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_products()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, 'Cliente', 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, 'Factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, 'Fecha de factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, 'Subtotal', 'header_right')
                            self._write_cell(worksheet, row, col+6, 'Descuento factura electrónica', 'header_right')
                            self._write_cell(worksheet, row, col+7, 'Subtotal sin IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+8, 'IVA', 'header_right')
                            self._write_cell(worksheet, row, col+9, 'IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+10, 'Total firmado', 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, 'Estado Tributación', 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount')
                                total_tax += line.get('tax_amount')
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            elif self.report_for == 'sales_team':

                self._write_cell(worksheet, row, col, 'Equipo de ventas', 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_sales_team()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, 'Cliente', 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, 'Factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, 'Fecha de factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, 'Subtotal', 'header_right')
                            self._write_cell(worksheet, row, col+6, 'Descuento factura electrónica', 'header_right')
                            self._write_cell(worksheet, row, col+7, 'Subtotal sin IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+8, 'IVA', 'header_right')
                            self._write_cell(worksheet, row, col+9, 'IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+10, 'Total firmado', 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, 'Estado Tributación', 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount')
                                total_tax += line.get('tax_amount')
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            elif self.report_for == 'sales_person':

                self._write_cell(worksheet, row, col, 'Vendedor', 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_sales_person()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, 'Cliente', 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, 'Factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, 'Fecha de factura', 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, 'Subtotal', 'header_right')
                            self._write_cell(worksheet, row, col+6, 'Descuento factura electrónica', 'header_right')
                            self._write_cell(worksheet, row, col+7, 'Subtotal sin IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+8, 'IVA', 'header_right')
                            self._write_cell(worksheet, row, col+9, 'IVA devuelto', 'header_right')
                            self._write_cell(worksheet, row, col+10, 'Total firmado', 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, 'Estado Tributación', 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, _('Total:'), 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount')
                                total_tax += line.get('tax_amount')
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            elif self.report_for == 'taxes':

                self._write_cell(worksheet, row, col, _('Impuesto'), 'header_center_bold')
                self._write_cell(worksheet, row, col+1, 'Total sin impuestos', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, 'Total impuesto cobrado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, 'Total impuesto pagado', 'header_right') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+4, 'Saldo', 'header_right') if not self.detailed_report else None
                row+=1

                data = self.get_report_datas_taxes()
                for rec in data:
                    self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                    self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+2, rec.get('total_tax_receievd'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+3, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                    self._write_cell(worksheet, row, col+4, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                    row+=1
                    for tex_line in rec.get('tax_groups'):
                        if self.detailed_report:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                            row+=1
                            # Headers for new amount columns (reordered)
                            self._write_cell(worksheet, row, col+2, _("Cliente"), 'header_center_bold')
                            self._write_cell(worksheet, row, col+3, _("Factura"), 'header_center_bold')
                            self._write_cell(worksheet, row, col+4, _("Fecha de factura"), 'header_center_bold')
                            self._write_cell(worksheet, row, col+5, _("Subtotal"), 'header_right')
                            self._write_cell(worksheet, row, col+6, _("Descuento factura electrónica"), 'header_right')
                            self._write_cell(worksheet, row, col+7, _("Subtotal sin IVA devuelto"), 'header_right')
                            self._write_cell(worksheet, row, col+8, _("IVA"), 'header_right')
                            self._write_cell(worksheet, row, col+9, _("IVA devuelto"), 'header_right')
                            self._write_cell(worksheet, row, col+10, _("Total firmado"), 'header_right')
                            # self._write_cell(worksheet, row, col+11, 'Sin impuestos', 'header_right')  # Hidden
                            # self._write_cell(worksheet, row, col+12, 'Sin impuestos firmado en moneda', 'header_right')  # Hidden
                            # Hidden columns: Monto pagado, Monto residual, Monto residual Signed, Impuesto, Impuesto de factura electrónica, Total, Total firmado en moneda
                            self._write_cell(worksheet, row, col+11, _("Estado Tributación"), 'header_center_bold')
                            row+=1
                            # Initialize totals
                            totals = {
                                'sub_total': 0,
                                'amount_discount_electronic_invoice': 0,
                                'amount_iva_returned': 0,
                                'amount_paid': 0,
                                'amount_residual': 0,
                                'amount_residual_signed': 0,
                                'amount_subtotal_without_iva_returned': 0,
                                'amount_tax': 0,
                                'amount_tax_electronic_invoice': 0,
                                'amount_tax_signed': 0,
                                'amount_total': 0,
                                'amount_total_in_currency_signed': 0,
                                'amount_total_signed': 0,
                                'amount_untaxed': 0,
                                'amount_untaxed_in_currency_signed': 0,
                            }
                            for line in tex_line.get('lines'):
                                self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                                self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                                self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                                self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                                self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                                self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                                self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                                self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                                # self._write_cell(worksheet, row, col+11, line.get('amount_untaxed', 0), 'right')  # Hidden
                                # self._write_cell(worksheet, row, col+12, line.get('amount_untaxed_in_currency_signed', 0), 'right')  # Hidden
                                self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')
                                # Update totals
                                for key in totals.keys():
                                    totals[key] += line.get(key, 0)
                                row+=1
                            # Write totals row (reordered)
                            self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                            self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                            self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                            self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                            self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                            self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                            # self._write_cell(worksheet, row, col+11, totals['amount_untaxed'], 'right_bold')  # Hidden
                            # self._write_cell(worksheet, row, col+12, totals['amount_untaxed_in_currency_signed'], 'right_bold')  # Hidden
                            row+=1
                        else:
                            self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                            total_base = 0
                            total_tax = 0
                            for line in tex_line.get('lines'):
                                total_base += line.get('base_amount',0) or 0
                                total_tax += line.get('tax_amount',0) or 0
                            self._write_cell(worksheet, row, col+2, total_base, 'right')
                            self._write_cell(worksheet, row, col+3, total_tax, 'right')
                            row+=1
                    row+=1

            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)

            export_id = self.env['account.excel.tax.report'].create({
                'excel_file': base64.b64encode(fp.getvalue()),
                'file_name': 'Reporte_Impuestos.xlsx',
            })
            fp.close()

            return {
                'type': 'ir.actions.act_window',
                'res_id': export_id.id,
                'res_model': 'account.excel.tax.report',
                'view_mode': 'form',
                'target': 'new',
                }

    def print_pdf_report(self):
        data = {}
        data['form'] = self.read(['start_date', 'end_date', 'tax_type', 'report_for', 'company_ids', 'detailed_report'])[0]

        if self.report_for == 'partners':
            data['report_data'] = self.get_report_datas_partners()
        elif self.report_for == 'productCategory':
            data['report_data'] = self.get_report_datas_productCategories()
        elif self.report_for == 'products':
            data['report_data'] = self.get_report_datas_products()
        elif self.report_for == 'sales_team':
            data['report_data'] = self.get_report_datas_sales_team()
        elif self.report_for == 'sales_person':
            data['report_data'] = self.get_report_datas_sales_person()
        elif self.report_for == 'taxes':
            data['report_data'] = self.get_report_datas_taxes()

        return self.env.ref('account_tax_report_omax.action_report_tax_report_account').report_action(self, data=data)

    def print_report(self):
        self.ensure_one()
        if self.report_format == 'excel':
            return self.print_xlsx_report()
        elif self.report_format == 'pdf':
            return self.print_pdf_report()
