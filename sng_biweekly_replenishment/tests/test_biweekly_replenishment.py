# -*- coding: utf-8 -*-

from datetime import timedelta
from io import BytesIO

from openpyxl import load_workbook

from odoo import Command, fields
from odoo.exceptions import AccessError
from odoo.tests import TransactionCase, tagged
from odoo.tests.common import new_test_user


@tagged("post_install", "-at_install")
class TestBiweeklyReplenishment(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.main_warehouse = cls.env["stock.warehouse"].create(
            {
                "name": "Bodega Principal de Prueba",
                "code": "CT00",
                "company_id": cls.env.company.id,
            }
        )
        cls.source_1 = cls.env["stock.warehouse"].create(
            {"name": "CEDIS Prioridad 1", "code": "CT01", "company_id": cls.env.company.id}
        )
        cls.source_2 = cls.env["stock.warehouse"].create(
            {"name": "CEDIS Prioridad 2", "code": "CT02", "company_id": cls.env.company.id}
        )
        cls.source_3 = cls.env["stock.warehouse"].create(
            {"name": "CEDIS Prioridad 3", "code": "CT03", "company_id": cls.env.company.id}
        )
        cls.group = cls.env["sng.warehouse.group"].create(
            {
                "name": "Grupo prueba reabastecimiento",
                "warehouse_ids": [
                    Command.set(
                        (
                            cls.main_warehouse
                            | cls.source_1
                            | cls.source_2
                            | cls.source_3
                        ).ids
                    )
                ],
            }
        )
        cls.product = cls.env["product.product"].create(
            {
                "name": "SKU prueba bisemanal",
                "default_code": "RBS-001",
                "is_storable": True,
            }
        )
        cls.planner = new_test_user(
            cls.env,
            login="replenishment_planner_test",
            name="Planificador de prueba",
            groups="sng_biweekly_replenishment.group_replenishment_planner",
        )
        cls.config = cls.env["sng.biweekly.replenishment.config"].create(
            {
                "name": "Configuración de prueba",
                "company_id": cls.env.company.id,
                "warehouse_group_id": cls.group.id,
                "main_warehouse_id": cls.main_warehouse.id,
                "demand_picking_type_ids": [Command.set(cls.main_warehouse.out_type_id.ids)],
                "product_ids": [Command.set(cls.product.ids)],
                "planner_user_ids": [Command.set(cls.planner.ids)],
                "coverage_days": 14,
                "safety_days": 2,
                "lead_time_days": 1,
                "source_line_ids": [
                    Command.create(
                        {
                            "sequence": 10,
                            "warehouse_id": cls.source_1.id,
                            "responsible_user_ids": [Command.set(cls.planner.ids)],
                        }
                    ),
                    Command.create({"sequence": 20, "warehouse_id": cls.source_2.id}),
                    Command.create({"sequence": 30, "warehouse_id": cls.source_3.id}),
                ],
            }
        )
        Quant = cls.env["stock.quant"]
        Quant._update_available_quantity(
            cls.product, cls.main_warehouse.lot_stock_id, 180.0
        )
        Quant._update_available_quantity(cls.product, cls.source_1.lot_stock_id, 70.0)
        Quant._update_available_quantity(cls.product, cls.source_2.lot_stock_id, 30.0)
        Quant._update_available_quantity(cls.product, cls.source_3.lot_stock_id, 50.0)
        cls.demand_move = cls._create_done_move(
            cls.product,
            140.0,
            cls.main_warehouse.lot_stock_id,
            cls.env.ref("stock.stock_location_customers"),
            cls.main_warehouse.out_type_id,
        )

    @classmethod
    def _create_done_move(
        cls, product, quantity, source, destination, picking_type, uom=None
    ):
        uom = uom or product.uom_id
        move = cls.env["stock.move"].create(
            {
                "name": product.display_name,
                "product_id": product.id,
                "product_uom_qty": quantity,
                "product_uom": uom.id,
                "location_id": source.id,
                "location_dest_id": destination.id,
                "picking_type_id": picking_type.id,
                "company_id": cls.env.company.id,
            }
        )
        move._action_confirm()
        move._action_assign()
        move.quantity = quantity
        move.picked = True
        move._action_done()
        move.date = fields.Datetime.now() - timedelta(days=1)
        return move

    def _new_batch(self, day_offset=0):
        batch = self.config._prepare_cycle(
            run_date=fields.Date.context_today(self.config) + timedelta(days=day_offset)
        )
        batch.action_recalculate()
        return batch

    def test_demand_scope_includes_external_outbound_move(self):
        period_end = fields.Datetime.now()
        period_start = period_end - timedelta(days=14)
        self.assertEqual(
            self.demand_move.location_id.warehouse_id,
            self.main_warehouse,
        )
        self.assertEqual(self.demand_move.state, "done")
        self.assertGreaterEqual(self.demand_move.date, period_start)
        self.assertLess(self.demand_move.date, period_end)
        self.assertEqual(self.demand_move.product_id, self.product)
        self.assertEqual(
            self.demand_move.picking_type_id,
            self.main_warehouse.out_type_id,
        )
        self.assertNotEqual(self.demand_move.location_dest_id.usage, "inventory")
        self.assertFalse(self.demand_move.is_inventory)
        base_domain = [
            ("state", "=", "done"),
            ("date", ">=", period_start),
            ("date", "<", period_end),
            ("product_id", "=", self.product.id),
            ("picking_type_id", "=", self.main_warehouse.out_type_id.id),
            ("location_id.warehouse_id", "=", self.main_warehouse.id),
            ("location_dest_id.usage", "!=", "inventory"),
            ("is_inventory", "=", False),
        ]
        self.assertIn(self.demand_move, self.env["stock.move"].search(base_domain))
        demand = self.config._get_demand_by_product(period_start, period_end)
        self.assertAlmostEqual(demand.get(self.product.id, 0.0), 140.0)

    def test_demand_window_boundaries_and_uom_conversion(self):
        self.env["stock.quant"]._update_available_quantity(
            self.product, self.main_warehouse.lot_stock_id, 100.0
        )
        period_end = fields.Datetime.now().replace(microsecond=0)
        period_start = period_end - timedelta(days=14)
        destination = self.env.ref("stock.stock_location_customers")
        at_start = self._create_done_move(
            self.product,
            10.0,
            self.main_warehouse.lot_stock_id,
            destination,
            self.main_warehouse.out_type_id,
        )
        before_start = self._create_done_move(
            self.product,
            20.0,
            self.main_warehouse.lot_stock_id,
            destination,
            self.main_warehouse.out_type_id,
        )
        at_end = self._create_done_move(
            self.product,
            30.0,
            self.main_warehouse.lot_stock_id,
            destination,
            self.main_warehouse.out_type_id,
        )
        dozens = self._create_done_move(
            self.product,
            2.0,
            self.main_warehouse.lot_stock_id,
            destination,
            self.main_warehouse.out_type_id,
            uom=self.env.ref("uom.product_uom_dozen"),
        )
        at_start.date = period_start
        before_start.date = period_start - timedelta(seconds=1)
        at_end.date = period_end
        dozens.date = period_start + timedelta(days=1)

        demand = self.config._get_demand_by_product(period_start, period_end)
        self.assertAlmostEqual(demand[self.product.id], 174.0)

    def test_formula_and_priority_allocation(self):
        batch = self._new_batch()
        line = batch.line_ids
        self.assertEqual(len(line), 1)
        self.assertAlmostEqual(line.demand_qty, 140.0)
        self.assertAlmostEqual(line.daily_demand, 10.0)
        self.assertAlmostEqual(line.target_stock, 160.0)
        self.assertAlmostEqual(line.reorder_point, 30.0)
        self.assertAlmostEqual(line.projected_qty, 40.0)
        self.assertAlmostEqual(line.suggested_qty, 120.0)
        self.assertEqual(
            line.allocation_ids.sorted("priority").mapped("allocated_qty"),
            [70.0, 30.0, 20.0],
        )
        self.assertAlmostEqual(line.shortage_qty, 0.0)

    def test_pickings_are_draft_and_idempotent(self):
        batch = self._new_batch()
        batch.action_generate_pickings()
        self.assertEqual(len(batch.picking_ids), 3)
        self.assertEqual(set(batch.picking_ids.mapped("state")), {"draft"})
        self.assertEqual(
            batch.picking_ids.mapped("picking_type_id"),
            (self.source_1 | self.source_2 | self.source_3).mapped("int_type_id"),
        )
        self.assertTrue(
            all(
                picking.location_dest_id == self.main_warehouse.lot_stock_id
                for picking in batch.picking_ids
            )
        )
        same_batch = self.config._prepare_cycle(run_date=batch.run_date)
        same_batch.action_generate_pickings()
        self.assertEqual(same_batch, batch)
        self.assertEqual(len(batch.picking_ids), 3)

    def test_previous_drafts_reduce_next_cycle(self):
        first_batch = self._new_batch()
        first_batch.action_generate_pickings()
        next_batch = self._new_batch(day_offset=1)
        line = next_batch.line_ids
        self.assertAlmostEqual(line.draft_in_qty, 120.0)
        self.assertAlmostEqual(line.projected_qty, 160.0)
        self.assertAlmostEqual(line.suggested_qty, 0.0)
        self.assertFalse(next_batch.allocation_ids)

    def test_shortage_and_cancellation_keep_history(self):
        self.env["stock.quant"]._update_available_quantity(
            self.product, self.source_3.lot_stock_id, -50.0
        )
        batch = self._new_batch()
        line = batch.line_ids
        self.assertAlmostEqual(line.allocated_qty, 100.0)
        self.assertAlmostEqual(line.shortage_qty, 20.0)
        batch.action_generate_pickings()
        self.assertEqual(batch.state, "partial")
        self.assertEqual(len(batch.picking_ids), 2)

        picking_ids = batch.picking_ids.ids
        batch.action_cancel()
        self.assertEqual(batch.state, "cancel")
        self.assertEqual(batch.picking_ids.ids, picking_ids)
        self.assertEqual(set(batch.picking_ids.mapped("state")), {"cancel"})

    def test_alert_lifecycle_does_not_duplicate_activity(self):
        self.env["stock.quant"]._update_available_quantity(
            self.product, self.main_warehouse.lot_stock_id, -15.0
        )
        self.config._check_reorder_alerts()
        alert = self.config.alert_ids
        self.assertEqual(alert.state, "open")
        self.assertAlmostEqual(alert.free_qty, 25.0)
        activity_count = len(alert.activity_ids.filtered("active"))
        self.assertEqual(activity_count, 1)

        self.config._check_reorder_alerts()
        self.assertEqual(len(alert.activity_ids.filtered("active")), activity_count)

        self.env["stock.quant"]._update_available_quantity(
            self.product, self.main_warehouse.lot_stock_id, 10.0
        )
        self.config._check_reorder_alerts()
        self.assertEqual(alert.state, "recovered")
        self.assertFalse(alert.activity_ids.filtered("active"))

        self.env["stock.quant"]._update_available_quantity(
            self.product, self.main_warehouse.lot_stock_id, -10.0
        )
        self.config._check_reorder_alerts()
        self.assertEqual(alert.state, "open")
        self.assertEqual(alert.open_count, 2)
        self.assertEqual(len(alert.activity_ids.filtered("active")), 1)

    def test_reports_render(self):
        batch = self._new_batch()
        batch.action_generate_pickings()
        pdf, pdf_type = self.env["ir.actions.report"].with_context(
            force_report_rendering=True
        )._render_qweb_pdf(
            "sng_biweekly_replenishment.action_replenishment_batch_pdf",
            res_ids=batch.ids,
        )
        self.assertEqual(pdf_type, "pdf")
        self.assertTrue(pdf.startswith(b"%PDF"))

        xlsx, xlsx_type = self.env["ir.actions.report"]._render(
            "sng_biweekly_replenishment.action_replenishment_batch_xlsx",
            batch.ids,
            {},
        )
        self.assertEqual(xlsx_type, "xlsx")
        sheet = load_workbook(BytesIO(xlsx), read_only=True).active
        self.assertEqual(sheet.cell(4, 1).value, "Código")
        self.assertEqual(sheet.cell(5, 1).value, self.product.default_code)

    def test_planner_cannot_create_configuration(self):
        with self.assertRaises(AccessError):
            self.env["sng.biweekly.replenishment.config"].with_user(
                self.planner
            ).create(
                {
                    "name": "No permitido",
                    "company_id": self.env.company.id,
                    "warehouse_group_id": self.group.id,
                    "main_warehouse_id": self.main_warehouse.id,
                }
            )
